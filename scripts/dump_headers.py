"""Imprime todos los headers (fila 1) de una hoja Excel especifica."""
from __future__ import annotations
import sys, io
from pathlib import Path
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

def main(path: str, sheet: str, n_rows: int = 5) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    print(f"Hoja: {sheet}  max_row={ws.max_row}  max_col={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=n_rows, values_only=True)):
        print(f"\n--- fila {i+1} ---")
        for col_idx, val in enumerate(row, start=1):
            if val is not None and str(val).strip() != "":
                print(f"  col {col_idx:3}: {val}")
    wb.close()

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], int(sys.argv[3]) if len(sys.argv) > 3 else 5)
