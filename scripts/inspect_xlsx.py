"""Inspector rapido de Excel: lista hojas, dimensiones y primeras columnas."""
from __future__ import annotations
import sys, io
from pathlib import Path
import openpyxl

# Forzar stdout UTF-8 para no romper en consolas cp1252 con emojis
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

def _s(v):
    if v is None: return ""
    s = str(v)
    return s.encode("utf-8", errors="replace").decode("utf-8")

def inspect(path: Path) -> None:
    print(f"\n{'='*80}\n{_s(path.name)}\n  {_s(path.parent)}\n{'='*80}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR cargando: {e}")
        return
    for sh in wb.sheetnames:
        ws = wb[sh]
        print(f"  [Hoja] {_s(sh)}  -> max_row={ws.max_row}  max_col={ws.max_column}")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
            cells = [_s(c) for c in row[:12]]
            tail = "  ..." if ws.max_column > 12 else ""
            print(f"      fila{i+1}: {cells}{tail}")
    wb.close()

if __name__ == "__main__":
    for p in sys.argv[1:]:
        inspect(Path(p))
