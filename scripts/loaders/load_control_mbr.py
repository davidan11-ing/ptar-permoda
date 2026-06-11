"""
Loader: FORMATO MBR UNICO.xlsx (hoja OPERACIONAL 2026) → condiciones_mbr_turno

El archivo tiene 1 fila por día (sin turno). Se inserta con turno=2 (mañana)
como convención para datos históricos diarios.

Columnas (0-based, header en fila 2 del Excel = row index 1):
  0: FECHA
  1: CAUDAL MBR 1 → mbr1_caudal_permeado
  2: CAUDAL MBR 2 → mbr2_caudal_permeado
  3: NIVEL MBR 1  (no existe en BD, se omite)
  4: NIVEL MBR 2  (no existe en BD, se omite)
  5: TMP MBR1     → mbr1_tmp
  6: TMP MBR 2    → mbr2_tmp
  7: PURGA        → mbr1_purga / mbr2_purga
  8: RECIRCULACIÓN → mbr1_recirculacion / mbr2_recirculacion
  9: OBSERVACIONES → observaciones

Valores N/R = null (no registrado)
Valores SI/NO = 1/0
"""

import openpyxl
import pymysql
import os
import warnings
from datetime import datetime, date
warnings.filterwarnings('ignore')

DB = dict(host='127.0.0.1', port=3306, user='root',
          password='Lsop5367**', database='ptar_permoda',
          charset='utf8mb4')

FILE = (r"C:\Users\lunaop\OneDrive - PERMODA LTDA\Documentos\Claude"
        r"\Projects\App PTAR 2\BASE DE DATOS\BALANCES 2026"
        r"\control de operacion\FORMATO MBR UNICO.xlsx")

HOJA = "OPERACIONAL 2026"
USUARIO = 'loader_excel'
TURNO_HISTORICO = 2   # convención para datos diarios sin turno


def clean_float(val):
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip().upper()
        if s in ('N/R', '', '-', 'NA', 'N/A', '#DIV/0!'):
            return None
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return None
    try:
        f = float(val)
        return f if f == f else None
    except (ValueError, TypeError):
        return None


def clean_bool(val):
    """SI → True, NO → False, None → None."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ('SI', 'SÍ', 'S', '1', 'YES', 'TRUE'):
        return True
    if s in ('NO', 'N', '0', 'FALSE'):
        return False
    return None


def clean_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except Exception:
        return None


def main():
    print("=" * 60)
    print("CARGANDO MBR OPERACIONAL 2026 -> condiciones_mbr_turno")
    print("=" * 60)

    if not os.path.exists(FILE):
        print(f"ERROR: No se encuentra el archivo: {FILE}")
        return

    wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

    # Buscar la hoja
    ws = None
    for sh in wb.sheetnames:
        if '2026' in sh.upper() and 'OPERACIONAL' in sh.upper():
            ws = wb[sh]
            break
    if ws is None:
        print(f"ERROR: No se encontró hoja OPERACIONAL 2026. Hojas: {wb.sheetnames}")
        wb.close()
        return

    print(f"  Hoja: '{ws.title}'")

    # Header en fila 2 (row_idx 1), datos desde fila 3
    registros = []
    hoy = date.today()

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        fecha = clean_date(row[0])
        if fecha is None:
            continue
        if fecha > hoy:
            continue

        caudal1  = clean_float(row[1])
        caudal2  = clean_float(row[2])
        # row[3] = NIVEL MBR1, row[4] = NIVEL MBR2 — se omiten (no existen en BD)
        tmp1     = clean_float(row[5])
        tmp2     = clean_float(row[6])
        purga    = clean_bool(row[7])
        recir    = clean_bool(row[8])
        obs_raw  = row[9] if len(row) > 9 else None
        obs      = str(obs_raw).strip() if obs_raw and str(obs_raw).strip().upper() not in ('NONE', '', 'N/R') else None

        # Solo insertar si hay al menos un dato
        if all(v is None for v in [caudal1, caudal2, tmp1, tmp2, purga]):
            continue

        registros.append({
            'fecha':               fecha,
            'turno':               TURNO_HISTORICO,
            'usuario':             USUARIO,
            'mbr1_caudal_permeado': caudal1,
            'mbr1_tmp':            tmp1,
            'mbr1_purga':          1 if purga else 0 if purga is not None else None,
            'mbr1_recirculacion':  1 if recir else 0 if recir is not None else None,
            'mbr2_caudal_permeado': caudal2,
            'mbr2_tmp':            tmp2,
            'mbr2_purga':          1 if purga else 0 if purga is not None else None,
            'mbr2_recirculacion':  1 if recir else 0 if recir is not None else None,
            'observaciones':       obs,
        })

    wb.close()
    print(f"  Filas a insertar: {len(registros)}")

    if not registros:
        print("  Sin datos para cargar.")
        return

    conn = pymysql.connect(**DB)
    cursor = conn.cursor()
    inserted = updated = 0

    sql = """
        INSERT INTO condiciones_mbr_turno
            (fecha, turno, usuario,
             mbr1_caudal_permeado, mbr1_tmp, mbr1_purga, mbr1_recirculacion,
             mbr2_caudal_permeado, mbr2_tmp, mbr2_purga, mbr2_recirculacion,
             observaciones)
        VALUES
            (%s, %s, %s,
             %s, %s, %s, %s,
             %s, %s, %s, %s,
             %s)
        ON DUPLICATE KEY UPDATE
            usuario=%s,
            mbr1_caudal_permeado=%s, mbr1_tmp=%s,
            mbr1_purga=%s, mbr1_recirculacion=%s,
            mbr2_caudal_permeado=%s, mbr2_tmp=%s,
            mbr2_purga=%s, mbr2_recirculacion=%s,
            observaciones=%s
    """

    for r in registros:
        vals_ins = (
            r['fecha'], r['turno'], r['usuario'],
            r['mbr1_caudal_permeado'], r['mbr1_tmp'],
            r['mbr1_purga'], r['mbr1_recirculacion'],
            r['mbr2_caudal_permeado'], r['mbr2_tmp'],
            r['mbr2_purga'], r['mbr2_recirculacion'],
            r['observaciones'],
        )
        vals_upd = (
            r['usuario'],
            r['mbr1_caudal_permeado'], r['mbr1_tmp'],
            r['mbr1_purga'], r['mbr1_recirculacion'],
            r['mbr2_caudal_permeado'], r['mbr2_tmp'],
            r['mbr2_purga'], r['mbr2_recirculacion'],
            r['observaciones'],
        )
        cursor.execute(sql, vals_ins + vals_upd)
        if cursor.rowcount == 1:
            inserted += 1
        elif cursor.rowcount == 2:
            updated += 1

    conn.commit()
    cursor.close()
    conn.close()

    print("=" * 60)
    print(f"  Insertados:   {inserted}")
    print(f"  Actualizados: {updated}")
    print(f"  Total:        {inserted + updated}")
    print("=" * 60)


if __name__ == '__main__':
    main()
