"""Loader Calidad de Agua: lee la hoja 'Tabla datos 1' de los archivos
mensuales DASHBOARD CALIDAD AGUA <MES> 2026.xlsm y carga las mediciones
en ptar_permoda.medicion_calidad (formato LONG).

Estructura del Excel (Tabla datos 1):
    Col 1: DATO (id consecutivo del Excel — se ignora)
    Col 2: FECHA
    Col 3: PARAMETRO (texto descriptivo)
    Col 4: TURNO (1, 2, 3)
    Cols 5-19: 15 unidades de tratamiento (PULMON .. RO_RECHAZO)
    Col 20+:   COSTO y columnas adicionales (se ignoran aqui)

Cada celda (col 5..19) con valor numerico distinto de 0 se inserta como
una fila en medicion_calidad. Valores 0 / NULL / formula con error se
omiten (en este Excel '0' significa "no medido", no "concentracion cero").

Uso:
  python scripts/loaders/load_calidad_agua.py "ruta\\DASHBOARD CALIDAD AGUA ABRIL 2026.xlsm"

  # o multiple archivos:
  python scripts/loaders/load_calidad_agua.py "ENERO.xlsm" "FEBRERO.xlsm" ...
"""
from __future__ import annotations
import sys, argparse, shutil, tempfile
from datetime import datetime, date
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import clean_date  # noqa: E402
from db import get_engine      # noqa: E402

HOJA_TARGET = "Tabla datos 1"

# Mapeo nombre largo del Excel -> codigo del catalogo parametro_calidad
PARAM_MAP = {
    "Temperatura(°C)":                                "TEMP",
    "pH (Unidades de pH)":                            "PH",
    "Demanda química de oxígeno (DQO)(mg/L)":         "DQO",
    "SOLIDOS DISUELTOS TOTALES (TDS)(mg/L)":          "TDS",
    "Sólidos suspendidos Totales(mg/L)":              "SST",
    "Sólidos Sedimentables (mg/L)":                   "SS",
    "HIERRO(ml/L)":                                   "FE",
    "Solidos Suspendidos totales GRAVIMETRICO(mg/L)": "SST_GRAV",
    "Cloruros (mg/L)":                                "CL",
    "FOSFORO TOTAL(mg/L)":                            "P",
    "Nitrógeno Total(mg/L)":                          "N",
    "Sulfatos (mg/L)":                                "SO4",
    "Alcalinidad (mg CaCO3/L)":                       "ALC",
    "Dureza Cálcica(mg CaCO3/L)":                     "DUR_CA",
    "Dureza Total (mg CaCO3/L)":                      "DUR_TOT",
    "SILICE(mg/L)":                                   "SIO2",
    "ORP(-MV)":                                       "ORP",
    "CLORO RES(mg/L)":                                "CLR",
    "Conductividad(Us/cm)":                           "COND",
    "Color (UPTCO)":                                  "COLOR",
    "Turbidez(NTU)":                                  "TURB",
}

# Columna Excel (1-based) -> codigo de unidad_tratamiento
COL_A_UNIDAD = {
    5:  "PULMON",
    6:  "HOMO",
    7:  "GEM_SAL",
    8:  "ANOXICO",
    9:  "MBBR",
    10: "MBR1_INT",
    11: "MBR2_INT",
    12: "MBR1_PER",
    13: "MBR2_PER",
    14: "VERTIMIENTO",
    15: "RO1_COMP",
    16: "RO1_E1",
    17: "RO1_E2",
    18: "RO2_PER",
    19: "RO_RECHAZO",
}


def clean_valor(v) -> float | None:
    """Convierte celda Excel a float; '0' significa 'no medido' -> None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f != 0 else None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith('#') or s.startswith('='):
            return None
        # quitar separadores comunes
        s = s.replace(',', '.')
        try:
            f = float(s)
            return f if f != 0 else None
        except ValueError:
            return None
    return None


def clean_turno(v) -> int | None:
    if isinstance(v, (int, float)):
        t = int(v)
        return t if t in (1, 2, 3) else None
    if isinstance(v, str):
        s = v.strip()
        if s in ('1', '2', '3'):
            return int(s)
    return None


def cargar_catalogos(engine):
    """Devuelve dicts {codigo: id} para parametros y unidades."""
    with engine.connect() as conn:
        params = {r[0]: r[1] for r in conn.execute(text(
            "SELECT codigo, id FROM parametro_calidad"
        )).all()}
        unidades = {r[0]: r[1] for r in conn.execute(text(
            "SELECT codigo, id FROM unidad_tratamiento"
        )).all()}
    return params, unidades


def load_archivo(path: Path, engine, params: dict, unidades: dict) -> tuple[int, int]:
    """Procesa un archivo Excel y devuelve (filas_procesadas, mediciones_insertadas)."""
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    shutil.copy2(path, tmp)

    try:
        wb = openpyxl.load_workbook(str(tmp), read_only=True, data_only=True, keep_vba=False)

        ws = None
        for sh in wb.sheetnames:
            if HOJA_TARGET.lower() in sh.lower():
                ws = wb[sh]
                break
        if ws is None:
            print(f"  [WARN] Hoja '{HOJA_TARGET}' no encontrada en {path.name}")
            wb.close()
            return 0, 0

        # Pre-resolver mapeo col_excel -> unidad_id
        col_a_unidad_id: dict[int, int] = {}
        for c, codigo in COL_A_UNIDAD.items():
            if codigo not in unidades:
                print(f"  [WARN] Unidad '{codigo}' falta en catalogo.")
                continue
            col_a_unidad_id[c] = unidades[codigo]

        mediciones = []
        filas_procesadas = 0
        params_no_mapeados: set[str] = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) < 5:
                continue

            fecha = clean_date(row[1])
            parametro_txt = row[2]
            turno = clean_turno(row[3])

            if fecha is None or parametro_txt is None or turno is None:
                continue

            param_codigo = PARAM_MAP.get(str(parametro_txt).strip())
            if param_codigo is None:
                params_no_mapeados.add(str(parametro_txt).strip())
                continue
            param_id = params.get(param_codigo)
            if param_id is None:
                continue

            filas_procesadas += 1

            # Recorre cols 5..19 (15 unidades)
            for col_excel, unidad_id in col_a_unidad_id.items():
                idx = col_excel - 1
                if idx >= len(row):
                    continue
                v = clean_valor(row[idx])
                if v is None:
                    continue
                mediciones.append({
                    "fecha": fecha,
                    "turno": turno,
                    "parametro_id": param_id,
                    "unidad_id": unidad_id,
                    "valor": v,
                })

        wb.close()
        wb = None

        if params_no_mapeados:
            print(f"  [WARN] Parametros sin mapear: {params_no_mapeados}")

        if not mediciones:
            return filas_procesadas, 0

        # UPSERT: mismo (fecha, turno, parametro, unidad) sobreescribe valor.
        # No hay UNIQUE KEY definida sobre esos 4 campos, asi que primero
        # eliminamos las filas existentes del periodo+parametros cargados
        # y luego insertamos limpio.
        # Estrategia mas simple: borrar por (fecha, turno, parametro_id, unidad_id)
        # ANTES de cada batch usando un SELECT distinct y un DELETE.
        # Alternativa: agregar UNIQUE KEY. Optamos por agregar la KEY si no existe.
        with engine.begin() as conn:
            # Crear UNIQUE KEY si no existe (idempotente)
            try:
                conn.execute(text("""
                    ALTER TABLE medicion_calidad
                    ADD UNIQUE KEY uq_medicion (fecha, turno, parametro_id, unidad_id)
                """))
            except Exception:
                pass  # ya existe

        sql = text("""
            INSERT INTO medicion_calidad (fecha, turno, parametro_id, unidad_id, valor)
            VALUES (:fecha, :turno, :parametro_id, :unidad_id, :valor)
            ON DUPLICATE KEY UPDATE valor = VALUES(valor)
        """)

        BATCH = 500
        n_insert = 0
        for i in range(0, len(mediciones), BATCH):
            with engine.begin() as conn:
                conn.execute(sql, mediciones[i:i + BATCH])
            n_insert += len(mediciones[i:i + BATCH])

        return filas_procesadas, n_insert

    finally:
        tmp.unlink(missing_ok=True)


def main():
    ap = argparse.ArgumentParser(
        description="Carga mediciones de calidad de agua desde uno o varios "
                    "archivos DASHBOARD CALIDAD AGUA *.xlsm"
    )
    ap.add_argument("archivos", nargs='+', help="Uno o varios archivos .xlsm")
    args = ap.parse_args()

    engine = get_engine()
    params, unidades = cargar_catalogos(engine)
    print(f"\n[CALIDAD AGUA] catalogos: {len(params)} parametros, {len(unidades)} unidades")

    total_filas = 0
    total_mediciones = 0
    for archivo in args.archivos:
        path = Path(archivo)
        if not path.exists():
            print(f"\n[ERROR] No existe: {path}")
            continue
        print(f"\n[CALIDAD AGUA] {path.name}")
        f, n = load_archivo(path, engine, params, unidades)
        print(f"  Filas procesadas:  {f}")
        print(f"  Mediciones cargadas: {n}")
        total_filas += f
        total_mediciones += n

    # Stats finales
    with engine.connect() as conn:
        st = conn.execute(text("""
            SELECT COUNT(*)                                    AS total,
                   COUNT(DISTINCT fecha)                       AS fechas,
                   COUNT(DISTINCT parametro_id)                AS parametros,
                   COUNT(DISTINCT unidad_id)                   AS unidades,
                   MIN(fecha)                                  AS desde,
                   MAX(fecha)                                  AS hasta
            FROM medicion_calidad
        """)).one()

    print(f"\n==== RESUMEN ====")
    print(f"  Archivos procesados:   {len(args.archivos)}")
    print(f"  Filas leidas:          {total_filas}")
    print(f"  Mediciones cargadas:   {total_mediciones}")
    print(f"\n  Estado de medicion_calidad:")
    print(f"    Total mediciones:    {st[0]}")
    print(f"    Fechas distintas:    {st[1]}")
    print(f"    Parametros activos:  {st[2]}")
    print(f"    Unidades activas:    {st[3]}")
    print(f"    Rango:               {st[4]} -> {st[5]}")
    print("DONE.")


if __name__ == "__main__":
    main()
