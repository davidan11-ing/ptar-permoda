"""Loader Contadores PTAR 2026.xlsx -> ptar_permoda.contadores_lectura (formato ANCHO).

Carga la hoja 'Contadores Por Turno 2026':
  - Una fila de Excel = fecha + hora + 35 medidores (lecturas acumulativas BIGINT)
  - Turno derivado de la hora: 22:00->1 | 06:00->2 | 14:00->3
  - En el Excel el turno es mayormente 1 (solo lectura nocturna); filas extras
    con otras horas se cargan con el turno correspondiente.
  - Si la fecha esta vacia pero la hora si, se propaga la ultima fecha valida
    (el Excel agrupa 3 turnos bajo una sola fecha en col A).

Uso:
  python scripts/loaders/load_contadores.py "ruta\\Contadores PTAR 2026.xlsx"
"""
from __future__ import annotations
import sys, argparse
from datetime import datetime, date, time as dtime
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import clean_int, clean_date  # noqa: E402
from db import get_engine  # noqa: E402

HOJA_TARGET = "Contadores Por Turno"

# Mapa hora -> numero de turno
HORA_A_TURNO = {
    dtime(22, 0, 0): 1,
    dtime(22, 0): 1,
    dtime(6, 0, 0): 2,
    dtime(6, 0): 2,
    dtime(14, 0, 0): 3,
    dtime(14, 0): 3,
}

# Mapa turno -> hora canonica (inverso; usado para cumplir chk_hora_turno)
TURNO_A_HORA = {1: "22:00:00", 2: "06:00:00", 3: "14:00:00"}

# Mapa col_excel (1-based) -> nombre de campo en contadores_lectura
# Generado del catálogo medidor + mapeo Excel->SQL del contadores_ptar_mysql.sql
COL_A_CAMPO = {
    3:  'entrada_ap_principal_6in',
    4:  'entrada_ap_fria_lavanderia_4in',
    5:  'entrada_ap_lab_lavanderia',
    6:  'entrada_medidor_rojo_tintoreria_4in',
    7:  'entrada_ap_fria_tintoreria_4in',
    8:  'entrada_medidor_rojo_lavanderia_4in',
    9:  'rama',
    10: 'abridora_1',
    11: 'abridora_2',
    12: 'tanque_reuso_2in',
    13: 'ptar',
    14: 'entrada_ro1',
    15: 'salida_ro1',
    16: 'entrada_ro2',
    17: 'salida_ro2',
    18: 'entrada_ap_rotativa_3in',
    19: 'medidor_verde_retorno',
    20: 'entrada_ap_tintoreria_6in',       # 100% NULL en 2026 (activo=0)
    21: 'envio_th',
    22: 'mbr1',
    23: 'mbr2',
    24: 'ingreso_uf_ptap',
    25: 'salida_uf_ptap',
    26: 'entrada_ap_ptar2_acueducto',
    27: 'entrada_ap_puerta4_acueducto',
    28: 'entrada_ap_quimicos',
    29: 'agua_caliente_tintoreria',
    30: 'medidor_prueba_agua_caliente',
    31: 'entrada_ap_puerta2_acueducto',
    32: 'entrada_ap_caldera_acueducto',
    33: 'entrada_ap_puerta5_acueducto',
    34: 'entrada_ap_puerta6_acueducto',    # 100% NULL en 2026 (activo=0)
    35: 'entrada_ap_puerta7_acueducto',
    36: 'entrada_ap_lavanderia_acueducto',
    37: 'entrada_ap_zona_lodos_acueducto',
}

ALL_FIELDS = list(COL_A_CAMPO.values())


def find_sheet(wb, contains: str):
    for sh in wb.sheetnames:
        if contains.lower() in sh.lower():
            return wb[sh]
    raise KeyError(f"Hoja con '{contains}' no encontrada. Disponibles: {wb.sheetnames}")


def clean_hora(v) -> dtime | None:
    if isinstance(v, dtime):
        return v
    if isinstance(v, datetime):
        return v.time()
    return None


def hora_a_turno(h: dtime | None) -> int:
    """22:00->1, 06:00->2, 14:00->3. Si no reconoce, infiere por minutos."""
    if h is None:
        return 1  # default: turno noche (lectura mas comun en el Excel)
    t = HORA_A_TURNO.get(h)
    if t:
        return t
    # fallback por hora
    if h.hour == 22:
        return 1
    if h.hour == 6:
        return 2
    if h.hour == 14:
        return 3
    return 1  # default


def load_contadores(ws, engine) -> int:
    filas = []
    ultima_fecha = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        # col A (idx 0) = fecha; puede estar vacia si hay 3 turnos bajo la misma fecha
        f = clean_date(row[0])
        if f is not None:
            ultima_fecha = f
        else:
            f = ultima_fecha

        if f is None:
            continue

        hora = clean_hora(row[1])
        turno = hora_a_turno(hora)
        hora_str = TURNO_A_HORA[turno]   # hora canonica; evita artefacto 00:00 del Excel

        # Leer los 35 campos de medidores
        rec: dict = {
            "fecha": f,
            "turno": turno,
            "hora_lectura": hora_str,
        }
        tiene_valor = False
        for col_excel, campo in COL_A_CAMPO.items():
            idx = col_excel - 1
            v = row[idx] if idx < len(row) else None
            n = clean_int(v)
            rec[campo] = n
            if n is not None:
                tiene_valor = True

        if not tiene_valor:
            continue  # fila completamente vacia

        filas.append(rec)

    if not filas:
        return 0

    # Construir INSERT dinamico con todos los campos
    cols = ["fecha", "turno", "hora_lectura"] + ALL_FIELDS
    placeholders = ", ".join(f":{c}" for c in cols)
    col_list = ", ".join(cols)
    updates = ", ".join(f"{c}=VALUES({c})" for c in ALL_FIELDS)  # no actualiza fecha/turno

    sql = text(f"""
        INSERT INTO contadores_lectura ({col_list})
        VALUES ({placeholders})
        ON DUPLICATE KEY UPDATE {updates}
    """)

    BATCH = 200
    n = 0
    for i in range(0, len(filas), BATCH):
        with engine.begin() as conn:
            conn.execute(sql, filas[i:i + BATCH])
        n += len(filas[i:i + BATCH])

    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Ruta al Contadores PTAR 2026.xlsx")
    args = ap.parse_args()

    path = Path(args.xlsx)
    print(f"\n[CONTADORES] {path.name}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    engine = get_engine()

    ws = find_sheet(wb, HOJA_TARGET)
    print(f"  Hoja: '{ws.title}'  max_row={ws.max_row}")

    n = load_contadores(ws, engine)
    wb.close()

    # stats
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM contadores_lectura")).scalar()
        fechas = conn.execute(text("SELECT COUNT(DISTINCT fecha) FROM contadores_lectura")).scalar()
        rango = conn.execute(text("SELECT MIN(fecha), MAX(fecha) FROM contadores_lectura")).one()

    print(f"  Filas upserted: {n}")
    print(f"  Total en BD:    {total}  ({fechas} fechas | {rango[0]} → {rango[1]})")
    print("DONE.")


if __name__ == "__main__":
    main()
