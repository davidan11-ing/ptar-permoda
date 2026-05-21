"""Loader Balance Hídrico: extrae entradas manuales de la hoja 'BASE DE DATOS'
del archivo DASHBOARD BALANCE HIDRICO BOGOTA 2026.xlsx y las inserta en
ptar_permoda.balance_hidrico_manual.

Solo se cargan los campos que NO derivan de contadores:
  - carrotanques_m3   (col C7)
  - kg_tela           (col C31)
  - und_efectivas     (col C38)
  - m_tela            (col C43)
  - mulas_funza_m3    (col C50)

Los demás campos del balance hídrico (PTAP, RO, envío TH, etc.)
provienen de consumo_turno y se exponen a través de v_balance_hidrico.

Uso:
  python scripts/loaders/load_balance_hidrico.py ^
    "BASE DE DATOS\\BALANCES 2026\\BALANCE HIDRICO\\FORMULACIÓN BALANCE\\DASHBOARD BALANCE HIDRICO BOGOTA 2026.xlsx"
"""
from __future__ import annotations
import sys, argparse, shutil, os, tempfile
from datetime import datetime, date
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import clean_int, clean_date  # noqa: E402
from db import get_engine               # noqa: E402

HOJA_TARGET = "BASE DE DATOS"

# Columnas manuales: (col_excel_1based, campo_sql, tipo)
# tipo: 'num' → float/decimal, 'int' → entero
MANUAL_COLS = {
    7:  ("carrotanques_m3", "num"),
    31: ("kg_tela",         "num"),
    38: ("und_efectivas",   "num"),
    43: ("m_tela",          "num"),
    50: ("mulas_funza_m3",  "num"),
}

# Columnas de identidad
COL_FECHA = 1
COL_TURNO = 2


def clean_num(v) -> float | None:
    """Convierte celda de Excel a float; devuelve None si inválido."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v != 0 else None   # 0 → None para no ensuciar la tabla
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith('#') or s.startswith('='):
            return None
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return None
    return None


def clean_turno(v) -> int | None:
    """Extrae turno (1/2/3) de la celda; ignora valores fuera de rango."""
    if isinstance(v, (int, float)):
        t = int(v)
        return t if t in (1, 2, 3) else None
    if isinstance(v, str):
        s = v.strip()
        if s in ('1', '2', '3'):
            return int(s)
    return None


def load_balance_manual(ws, engine) -> tuple[int, int]:
    """Lee la hoja y carga los campos manuales en balance_hidrico_manual.

    Returns (filas_procesadas, filas_con_algun_valor_manual).
    """
    rows_data = []
    ultima_fecha: date | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        # ── Fecha ────────────────────────────────────────────────────
        raw_fecha = row[COL_FECHA - 1]
        f = clean_date(raw_fecha)
        if f is None and isinstance(raw_fecha, str) and raw_fecha.startswith('='):
            # formula de fecha (=+A312 etc.) → usar ultima_fecha
            f = ultima_fecha
        if f is not None:
            ultima_fecha = f
        if f is None:
            continue

        # ── Turno ────────────────────────────────────────────────────
        turno = clean_turno(row[COL_TURNO - 1])
        if turno is None:
            continue

        # ── Campos manuales ──────────────────────────────────────────
        rec = {"fecha": f, "turno": turno}
        tiene_manual = False
        for col_num, (campo, _) in MANUAL_COLS.items():
            idx = col_num - 1
            raw = row[idx] if idx < len(row) else None
            val = clean_num(raw)
            rec[campo] = val
            if val is not None:
                tiene_manual = True

        rows_data.append((tiene_manual, rec))

    filas_totales = len(rows_data)
    filas_con_manual = sum(1 for tm, _ in rows_data if tm)

    if not rows_data:
        return 0, 0

    # UPSERT — inserta TODO (incluyendo filas sin manuales) para que la
    # vista pueda hacer LEFT JOIN y mostrar todos los turnos.
    # Esto asegura que existan filas con carrotanques=0, mulas=0, etc.
    campos_manuales = [c for c, _ in MANUAL_COLS.values()]
    col_list = "fecha, turno, " + ", ".join(campos_manuales)
    placeholders = ":fecha, :turno, " + ", ".join(f":{c}" for c in campos_manuales)
    updates = ", ".join(
        f"{c}=COALESCE(VALUES({c}), {c})"   # no sobreescribe con NULL si ya hay valor
        for c in campos_manuales
    )

    sql = text(f"""
        INSERT INTO balance_hidrico_manual ({col_list})
        VALUES ({placeholders})
        ON DUPLICATE KEY UPDATE {updates}
    """)

    # Campos NOT NULL DEFAULT 0 — convertir None a 0.0
    NOT_NULL_ZERO = {"carrotanques_m3", "mulas_funza_m3"}

    records = []
    for _, rec in rows_data:
        r = dict(rec)
        for c in campos_manuales:
            if c not in r:
                r[c] = None
            # NOT NULL cols: None → 0.0
            if r[c] is None and c in NOT_NULL_ZERO:
                r[c] = 0.0
        records.append(r)

    BATCH = 300
    n = 0
    for i in range(0, len(records), BATCH):
        with engine.begin() as conn:
            conn.execute(sql, records[i:i + BATCH])
        n += len(records[i:i + BATCH])

    return filas_totales, filas_con_manual


def main():
    ap = argparse.ArgumentParser(
        description="Carga entradas manuales de balance hídrico desde el Dashboard Excel"
    )
    ap.add_argument("xlsx", help="Ruta al DASHBOARD BALANCE HIDRICO BOGOTA 2026.xlsx")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"[ERROR] Archivo no encontrado: {path}")
        sys.exit(1)

    print(f"\n[BALANCE HÍDRICO MANUAL] {path.name}")

    # Copiar a temp para no bloquear el archivo si está abierto en Excel
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    shutil.copy2(path, tmp)

    try:
        wb = openpyxl.load_workbook(str(tmp), read_only=True, data_only=True)
        print(f"  Hojas disponibles: {wb.sheetnames}")

        # Buscar la hoja BASE DE DATOS
        ws = None
        for sh in wb.sheetnames:
            if HOJA_TARGET.lower() in sh.lower():
                ws = wb[sh]
                break
        if ws is None:
            print(f"[ERROR] Hoja '{HOJA_TARGET}' no encontrada.")
            sys.exit(1)

        print(f"  Hoja: '{ws.title}'  max_row={ws.max_row}")

        engine = get_engine()
        total, con_manuales = load_balance_manual(ws, engine)
        wb.close()
        wb = None   # liberar handle antes del unlink

        print(f"  Filas procesadas:              {total}")
        print(f"  Filas con algún campo manual:  {con_manuales}")

        # Estadísticas finales
        with engine.connect() as conn:
            stats = conn.execute(text("""
                SELECT
                    COUNT(*)                          AS total_filas,
                    COUNT(kg_tela)                    AS dias_con_kg_tela,
                    COUNT(und_efectivas)              AS dias_con_und,
                    COUNT(m_tela)                     AS dias_con_m_tela,
                    SUM(COALESCE(carrotanques_m3, 0)) AS m3_carrotanques_total,
                    SUM(COALESCE(mulas_funza_m3,  0)) AS m3_mulas_total,
                    MIN(fecha)                        AS desde,
                    MAX(fecha)                        AS hasta
                FROM balance_hidrico_manual
            """)).one()

        print(f"\n  En BD:")
        print(f"    Total filas:          {stats[0]}")
        print(f"    Turnos con kg_tela:   {stats[1]}")
        print(f"    Turnos con und_ef:    {stats[2]}")
        print(f"    Turnos con m_tela:    {stats[3]}")
        print(f"    Total carrotanques:   {stats[4]:.0f} m³")
        print(f"    Total mulas Funza:    {stats[5]:.0f} m³")
        print(f"    Rango:                {stats[6]} → {stats[7]}")
        print("DONE.")

    finally:
        tmp.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
