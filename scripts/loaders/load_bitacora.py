"""Loader BITACORA OPERACION PTAR (mensual) -> ptar_permoda.

Carga 2 hojas del archivo "FORMATO BITACORA OPERACION PTAR - <MES> 2026.xlsx":
  * INVENTARIO Y CONSUMO GEM    -> operacion_gem_turno    (3 turnos/dia)
  * REGISTRO RO                 -> operacion_ro_turno     (3 turnos/dia)

Los agregados mensuales de consumo químico se obtienen via vistas
v_consumo_quimico_mensual y v_quimico_real_vs_proyectado (no tabla directa).

Uso:
  python scripts/loaders/load_bitacora.py "ruta\\archivo.xlsx"
"""
from __future__ import annotations
import sys, argparse
from pathlib import Path

import openpyxl
from sqlalchemy import text

# importar common (ajusta sys.path) y luego db
sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import clean_num, clean_int, clean_str, clean_date  # noqa: E402
from db import get_engine  # noqa: E402


TURNO_MAP = {"1 (NOCHE)": 1, "2 (MAÑANA)": 2, "3 (TARDE)": 3,
             "1 (NOCHE)\n": 1, "2 (MANANA)": 2, "3 (TARDE)": 3}


def turno_num(v) -> int | None:
    """Numero de turno del dia (1=Noche, 2=Manana, 3=Tarde) a partir del texto descriptivo."""
    if v is None:
        return None
    s = str(v).strip().upper().replace("Ñ", "N")
    if "NOCHE" in s: return 1
    if "MANANA" in s or "MAÑANA" in s: return 2
    if "TARDE" in s: return 3
    if s in {"1", "2", "3"}:
        return int(s)
    return None


def turno_from_secuencial(num_turno_secuencial) -> int | None:
    """Convierte el contador '# TURNO' (1..93) al turno del dia (1..3)."""
    n = clean_int(num_turno_secuencial)
    if n is None or n < 1:
        return None
    return ((n - 1) % 3) + 1


def find_sheet(wb, contains: str):
    for sh in wb.sheetnames:
        if contains.lower() in sh.lower():
            return wb[sh]
    raise KeyError(f"No se encontro hoja que contenga '{contains}'. Disponibles: {wb.sheetnames}")


def load_gem(ws, engine) -> int:
    """Carga hoja INVENTARIO Y CONSUMO GEM. Retorna filas insertadas/actualizadas."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        fecha = clean_date(row[0])
        if fecha is None:
            continue
        # turno robusto: prioriza el texto descriptivo (NOCHE/MAÑANA/TARDE),
        # cae a derivar del contador secuencial '# TURNO' si el texto no es claro
        turno = turno_num(row[2]) or turno_from_secuencial(row[3])
        rec = {
            "fecha": fecha,
            "dia_mes": clean_int(row[1]),
            "turno": turno,
            "turno_descripcion": clean_str(row[2]),
            "diligencia_bitacora": clean_str(row[4]),
            "final_acido_l": clean_num(row[5]),
            "final_coagulante_l": clean_num(row[6]),
            "final_decolorante_l": clean_num(row[7]),
            "final_pol_anionico_kg": clean_num(row[8]),
            "final_pol_cationico_kg": clean_num(row[9]),
            "horometro_inicial": clean_int(row[10]),
            "caudal_total_tratado_gem_m3": clean_num(row[11]),
            "caudal_tratamiento_m3h": clean_num(row[12]),
            "consumo_acido_l": clean_num(row[13]),
            "consumo_coagulante_l": clean_num(row[14]),
            "consumo_decolorante_l": clean_num(row[15]),
            "consumo_pol_anionico_kg": clean_num(row[16]),
            "consumo_pol_cationico_kg": clean_num(row[17]),
            "kg_acido": clean_num(row[18]),
            "kg_coagulante": clean_num(row[19]),
            "kg_decolorante": clean_num(row[20]),
            "kg_pol_anionico": clean_num(row[21]),
            "kg_pol_cationico": clean_num(row[22]),
            "ppm_acido": clean_num(row[23]),
            "ppm_coagulante": clean_num(row[24]),
            "ppm_decolorante": clean_num(row[25]),
            "ppm_pol_anionico": clean_num(row[26]),
            "ppm_pol_cationico": clean_num(row[27]),
            # 28-32 densidades y 33-37 precios -> catalogo, omitidos
            "costo_op_acido": clean_num(row[38]),
            "costo_op_coagulante": clean_num(row[39]),
            "costo_op_decolorante": clean_num(row[40]),
            "costo_op_anionico": clean_num(row[41]),
            "costo_op_cationico": clean_num(row[42]),
            "costo_quimica_turno": clean_num(row[43]),
            "limite_indicador_m3": clean_num(row[44]),
            "pesos_por_m3": clean_num(row[45]),
        }
        if rec["turno"] is None:
            continue
        rows.append(rec)

    if not rows:
        return 0

    sql = text("""
        INSERT INTO operacion_gem_turno
          (fecha, dia_mes, turno, turno_descripcion, diligencia_bitacora,
           final_acido_l, final_coagulante_l, final_decolorante_l,
           final_pol_anionico_kg, final_pol_cationico_kg,
           horometro_inicial, caudal_total_tratado_gem_m3, caudal_tratamiento_m3h,
           consumo_acido_l, consumo_coagulante_l, consumo_decolorante_l,
           consumo_pol_anionico_kg, consumo_pol_cationico_kg,
           kg_acido, kg_coagulante, kg_decolorante, kg_pol_anionico, kg_pol_cationico,
           ppm_acido, ppm_coagulante, ppm_decolorante, ppm_pol_anionico, ppm_pol_cationico,
           costo_op_acido, costo_op_coagulante, costo_op_decolorante,
           costo_op_anionico, costo_op_cationico,
           costo_quimica_turno, limite_indicador_m3, pesos_por_m3)
        VALUES
          (:fecha, :dia_mes, :turno, :turno_descripcion, :diligencia_bitacora,
           :final_acido_l, :final_coagulante_l, :final_decolorante_l,
           :final_pol_anionico_kg, :final_pol_cationico_kg,
           :horometro_inicial, :caudal_total_tratado_gem_m3, :caudal_tratamiento_m3h,
           :consumo_acido_l, :consumo_coagulante_l, :consumo_decolorante_l,
           :consumo_pol_anionico_kg, :consumo_pol_cationico_kg,
           :kg_acido, :kg_coagulante, :kg_decolorante, :kg_pol_anionico, :kg_pol_cationico,
           :ppm_acido, :ppm_coagulante, :ppm_decolorante, :ppm_pol_anionico, :ppm_pol_cationico,
           :costo_op_acido, :costo_op_coagulante, :costo_op_decolorante,
           :costo_op_anionico, :costo_op_cationico,
           :costo_quimica_turno, :limite_indicador_m3, :pesos_por_m3)
        ON DUPLICATE KEY UPDATE
           turno_descripcion=VALUES(turno_descripcion),
           diligencia_bitacora=VALUES(diligencia_bitacora),
           final_acido_l=VALUES(final_acido_l),
           final_coagulante_l=VALUES(final_coagulante_l),
           final_decolorante_l=VALUES(final_decolorante_l),
           final_pol_anionico_kg=VALUES(final_pol_anionico_kg),
           final_pol_cationico_kg=VALUES(final_pol_cationico_kg),
           horometro_inicial=VALUES(horometro_inicial),
           caudal_total_tratado_gem_m3=VALUES(caudal_total_tratado_gem_m3),
           caudal_tratamiento_m3h=VALUES(caudal_tratamiento_m3h),
           consumo_acido_l=VALUES(consumo_acido_l),
           consumo_coagulante_l=VALUES(consumo_coagulante_l),
           consumo_decolorante_l=VALUES(consumo_decolorante_l),
           consumo_pol_anionico_kg=VALUES(consumo_pol_anionico_kg),
           consumo_pol_cationico_kg=VALUES(consumo_pol_cationico_kg),
           kg_acido=VALUES(kg_acido),
           kg_coagulante=VALUES(kg_coagulante),
           kg_decolorante=VALUES(kg_decolorante),
           kg_pol_anionico=VALUES(kg_pol_anionico),
           kg_pol_cationico=VALUES(kg_pol_cationico),
           ppm_acido=VALUES(ppm_acido),
           ppm_coagulante=VALUES(ppm_coagulante),
           ppm_decolorante=VALUES(ppm_decolorante),
           ppm_pol_anionico=VALUES(ppm_pol_anionico),
           ppm_pol_cationico=VALUES(ppm_pol_cationico),
           costo_op_acido=VALUES(costo_op_acido),
           costo_op_coagulante=VALUES(costo_op_coagulante),
           costo_op_decolorante=VALUES(costo_op_decolorante),
           costo_op_anionico=VALUES(costo_op_anionico),
           costo_op_cationico=VALUES(costo_op_cationico),
           costo_quimica_turno=VALUES(costo_quimica_turno),
           limite_indicador_m3=VALUES(limite_indicador_m3),
           pesos_por_m3=VALUES(pesos_por_m3)
    """)
    with engine.begin() as conn:
        conn.execute(sql, rows)
    return len(rows)


def load_ro(ws, engine) -> int:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        fecha = clean_date(row[0])
        if fecha is None:
            continue
        t = turno_num(row[2])
        if t is None:
            continue
        rec = {
            "fecha": fecha,
            "dia_mes": clean_int(row[1]),
            "turno": t,
            "turno_descripcion": clean_str(row[2]),
            # nota: RO no tiene contador secuencial, solo turno descriptivo
            "aplic_hcl": clean_num(row[3]),
            "aplic_kuriverter": clean_num(row[4]),
            "aplic_vitec": clean_num(row[5]),
            "aplic_naoh": clean_num(row[6]),
            "aplic_bisulfito": clean_num(row[7]),
            "tiempo_operacion_min": clean_int(row[8]),
            "cm_hcl": clean_num(row[9]),
            "cm_ik220": clean_num(row[10]),
            "cm_vitec7000": clean_num(row[11]),
            "cm_naoh": clean_num(row[12]),
            "cm_bisulfito": clean_num(row[13]),
            "inv_l_hcl": clean_num(row[14]),
            "inv_l_kuriverter": clean_num(row[15]),
            "inv_l_vitec": clean_num(row[16]),
            "inv_l_naoh": clean_num(row[17]),
            "inv_l_bisulfito": clean_num(row[18]),
            "consumo_l_hcl": clean_num(row[19]),
            "consumo_l_kuriverter": clean_num(row[20]),
            "consumo_l_vitec": clean_num(row[21]),
            "consumo_l_naoh": clean_num(row[22]),
            "consumo_l_bisulfito": clean_num(row[23]),
            # 24-28 densidades -> catalogo
            "consumo_kg_hcl": clean_num(row[29]),
            "consumo_kg_kuriverter": clean_num(row[30]),
            "consumo_kg_vitec": clean_num(row[31]),
            "consumo_kg_naoh": clean_num(row[32]),
            "consumo_kg_bisulfito": clean_num(row[33]),
            "volumen_enviado_ro_m3": clean_num(row[34]),
            "ppm_hcl": clean_num(row[35]),
            "ppm_kuriverter": clean_num(row[36]),
            "ppm_vitec": clean_num(row[37]),
            "ppm_naoh": clean_num(row[38]),
            "ppm_bisulfito": clean_num(row[39]),
            # 40-44 precios -> catalogo
            "costo_op_hcl": clean_num(row[45]),
            "costo_op_kuriverter": clean_num(row[46]),
            "costo_op_vitec": clean_num(row[47]),
            "costo_op_naoh": clean_num(row[48]),
            "costo_op_bisulfito": clean_num(row[49]),
            "costo_quimica_turno": clean_num(row[50]),
            "limite_indicador_m3": clean_num(row[51]),
            "pesos_m3_enviado_ro": clean_num(row[52]),
            "pesos_m3_rechazo": clean_num(row[53]),
            "pesos_m3_permeado_ro": clean_num(row[54]),
        }
        rows.append(rec)
    if not rows:
        return 0

    sql = text("""
        INSERT INTO operacion_ro_turno
          (fecha, dia_mes, turno, turno_descripcion,
           aplic_hcl, aplic_kuriverter, aplic_vitec, aplic_naoh, aplic_bisulfito,
           tiempo_operacion_min,
           cm_hcl, cm_ik220, cm_vitec7000, cm_naoh, cm_bisulfito,
           inv_l_hcl, inv_l_kuriverter, inv_l_vitec, inv_l_naoh, inv_l_bisulfito,
           consumo_l_hcl, consumo_l_kuriverter, consumo_l_vitec, consumo_l_naoh, consumo_l_bisulfito,
           consumo_kg_hcl, consumo_kg_kuriverter, consumo_kg_vitec, consumo_kg_naoh, consumo_kg_bisulfito,
           volumen_enviado_ro_m3,
           ppm_hcl, ppm_kuriverter, ppm_vitec, ppm_naoh, ppm_bisulfito,
           costo_op_hcl, costo_op_kuriverter, costo_op_vitec, costo_op_naoh, costo_op_bisulfito,
           costo_quimica_turno, limite_indicador_m3,
           pesos_m3_enviado_ro, pesos_m3_rechazo, pesos_m3_permeado_ro)
        VALUES
          (:fecha, :dia_mes, :turno, :turno_descripcion,
           :aplic_hcl, :aplic_kuriverter, :aplic_vitec, :aplic_naoh, :aplic_bisulfito,
           :tiempo_operacion_min,
           :cm_hcl, :cm_ik220, :cm_vitec7000, :cm_naoh, :cm_bisulfito,
           :inv_l_hcl, :inv_l_kuriverter, :inv_l_vitec, :inv_l_naoh, :inv_l_bisulfito,
           :consumo_l_hcl, :consumo_l_kuriverter, :consumo_l_vitec, :consumo_l_naoh, :consumo_l_bisulfito,
           :consumo_kg_hcl, :consumo_kg_kuriverter, :consumo_kg_vitec, :consumo_kg_naoh, :consumo_kg_bisulfito,
           :volumen_enviado_ro_m3,
           :ppm_hcl, :ppm_kuriverter, :ppm_vitec, :ppm_naoh, :ppm_bisulfito,
           :costo_op_hcl, :costo_op_kuriverter, :costo_op_vitec, :costo_op_naoh, :costo_op_bisulfito,
           :costo_quimica_turno, :limite_indicador_m3,
           :pesos_m3_enviado_ro, :pesos_m3_rechazo, :pesos_m3_permeado_ro)
        ON DUPLICATE KEY UPDATE
           turno_descripcion=VALUES(turno_descripcion),
           aplic_hcl=VALUES(aplic_hcl), aplic_kuriverter=VALUES(aplic_kuriverter),
           aplic_vitec=VALUES(aplic_vitec), aplic_naoh=VALUES(aplic_naoh),
           aplic_bisulfito=VALUES(aplic_bisulfito),
           tiempo_operacion_min=VALUES(tiempo_operacion_min),
           cm_hcl=VALUES(cm_hcl), cm_ik220=VALUES(cm_ik220),
           cm_vitec7000=VALUES(cm_vitec7000), cm_naoh=VALUES(cm_naoh),
           cm_bisulfito=VALUES(cm_bisulfito),
           inv_l_hcl=VALUES(inv_l_hcl), inv_l_kuriverter=VALUES(inv_l_kuriverter),
           inv_l_vitec=VALUES(inv_l_vitec), inv_l_naoh=VALUES(inv_l_naoh),
           inv_l_bisulfito=VALUES(inv_l_bisulfito),
           consumo_l_hcl=VALUES(consumo_l_hcl), consumo_l_kuriverter=VALUES(consumo_l_kuriverter),
           consumo_l_vitec=VALUES(consumo_l_vitec), consumo_l_naoh=VALUES(consumo_l_naoh),
           consumo_l_bisulfito=VALUES(consumo_l_bisulfito),
           consumo_kg_hcl=VALUES(consumo_kg_hcl), consumo_kg_kuriverter=VALUES(consumo_kg_kuriverter),
           consumo_kg_vitec=VALUES(consumo_kg_vitec), consumo_kg_naoh=VALUES(consumo_kg_naoh),
           consumo_kg_bisulfito=VALUES(consumo_kg_bisulfito),
           volumen_enviado_ro_m3=VALUES(volumen_enviado_ro_m3),
           ppm_hcl=VALUES(ppm_hcl), ppm_kuriverter=VALUES(ppm_kuriverter),
           ppm_vitec=VALUES(ppm_vitec), ppm_naoh=VALUES(ppm_naoh),
           ppm_bisulfito=VALUES(ppm_bisulfito),
           costo_op_hcl=VALUES(costo_op_hcl), costo_op_kuriverter=VALUES(costo_op_kuriverter),
           costo_op_vitec=VALUES(costo_op_vitec), costo_op_naoh=VALUES(costo_op_naoh),
           costo_op_bisulfito=VALUES(costo_op_bisulfito),
           costo_quimica_turno=VALUES(costo_quimica_turno),
           limite_indicador_m3=VALUES(limite_indicador_m3),
           pesos_m3_enviado_ro=VALUES(pesos_m3_enviado_ro),
           pesos_m3_rechazo=VALUES(pesos_m3_rechazo),
           pesos_m3_permeado_ro=VALUES(pesos_m3_permeado_ro)
    """)
    with engine.begin() as conn:
        conn.execute(sql, rows)
    return len(rows)


def main():
    ap = argparse.ArgumentParser(
        description="Carga INVENTARIO Y CONSUMO GEM + REGISTRO RO desde la bitácora mensual."
    )
    ap.add_argument("xlsx", help="Ruta al FORMATO BITACORA OPERACION PTAR - <MES> <ANIO>.xlsx")
    args = ap.parse_args()

    path = Path(args.xlsx)
    print(f"\n[BITACORA] {path.name}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    engine = get_engine()

    ws_gem = find_sheet(wb, "INVENTARIO Y CONSUMO GEM")
    n_gem = load_gem(ws_gem, engine)
    print(f"  operacion_gem_turno : {n_gem} filas upserted")

    ws_ro = find_sheet(wb, "REGISTRO RO")
    n_ro = load_ro(ws_ro, engine)
    print(f"  operacion_ro_turno  : {n_ro} filas upserted")

    wb.close()
    print("DONE.")
    print("  (Consumo mensual disponible via v_consumo_quimico_mensual)")


if __name__ == "__main__":
    main()
