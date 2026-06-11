"""
Loader: OPERACIÓN RO1 sheet → condiciones_ro_turno
Lee CONTROL Y SEGUIMIENTO RO.xlsx (hoja OPERACIÓN RO1) y carga
presiones y caudales históricos. Todos los registros se asignan turno=2
(mañana) por convención — el archivo no tiene turno explícito.

Columnas (0-based, header en fila 1):
  0:  Fecha
  1:  HORA (ignorado para turno)
  2:  P Entrada E1 → p_entrada_e1
  3:  P Salida E1  → p_salida_e1
  4:  P Entrada E2 → p_entrada_e2
  5:  P Salida E2  → p_salida_e2
  6:  dP E1        (calculado — skip)
  7:  dP E2        (calculado — skip)
  8:  Q Entrada E1 (se omite: va en operacion_ro_turno, no en condiciones)
  9:  Q Salida E1  → q_permeado_e1
  10-11: calculados (skip)
  12: Flujo norm E1 → flujo_normalizado_e1
  14: Q Salida E2  → q_permeado_e2
  19: Q PERMEADO RO1 MEDIDO → q_rechazo_rotametro
"""

import openpyxl
import pymysql
import os
from datetime import datetime, date
import warnings
warnings.filterwarnings('ignore')

DB = dict(host='127.0.0.1', port=3306, user='root',
          password='Lsop5367**', database='ptar_permoda',
          charset='utf8mb4')

FILE = (r"C:\Users\lunaop\OneDrive - PERMODA LTDA\Documentos\Claude"
        r"\Projects\App PTAR 2\BASE DE DATOS\BALANCES 2026"
        r"\control de operacion\CONTROL Y SEGUIMIENTO RO.xlsx")

TURNO_CONVENCIONAL = 2   # turno mañana — convención para datos sin turno explícito
USUARIO = 'loader_excel'


def clean_float(val):
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if s in ('#DIV/0!', '', '-', 'N/A', '#VALUE!', '#REF!'):
            return None
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        f = float(val)
        return f if f == f else None   # NaN check
    return None


def clean_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except Exception:
        return None


def main():
    print("=" * 60)
    print("CARGANDO CONDICIONES RO -> condiciones_ro_turno")
    print("=" * 60)

    if not os.path.exists(FILE):
        print(f"ERROR: No se encontró: {FILE}")
        return

    wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

    # Buscar hoja OPERACIÓN RO1
    ws = None
    for sh in wb.sheetnames:
        sh_ascii = sh.encode('ascii', 'replace').decode()
        if 'OPERACI' in sh_ascii.upper() and 'RO1' in sh_ascii.upper():
            ws = wb[sh]
            break
    if ws is None:
        print(f"No se encontró hoja OPERACIÓN RO1. Hojas: {wb.sheetnames}")
        wb.close()
        return

    print(f"  Hoja: '{ws.title}'")

    registros = []
    hoy = date.today()

    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha = clean_date(row[0])
        if fecha is None:
            continue

        # Saltar si todos los valores numéricos son nulos o errores
        vals_raw = [row[i] for i in (2, 3, 4, 5, 9, 12, 14)]
        if all(clean_float(v) is None for v in vals_raw):
            continue

        rec = {
            'fecha':               fecha,
            'turno':               TURNO_CONVENCIONAL,
            'usuario':             USUARIO,
            'p_entrada_e1':        clean_float(row[2]),
            'p_salida_e1':         clean_float(row[3]),
            'p_entrada_e2':        clean_float(row[4]),
            'p_salida_e2':         clean_float(row[5]),
            'q_permeado_e1':       clean_float(row[9]),
            'flujo_normalizado_e1': clean_float(row[12]) if len(row) > 12 else None,
            'q_permeado_e2':       clean_float(row[14]) if len(row) > 14 else None,
            'q_rechazo_rotametro': clean_float(row[19]) if len(row) > 19 else None,
        }

        # Solo insertar si hay al menos una presión o caudal válido
        campos_util = ['p_entrada_e1', 'p_salida_e1', 'p_entrada_e2', 'p_salida_e2',
                       'q_permeado_e1', 'q_permeado_e2']
        if all(rec[c] is None for c in campos_util):
            continue

        registros.append(rec)

    wb.close()
    print(f"  Filas con datos: {len(registros)}")

    if not registros:
        print("  Sin datos para cargar.")
        return

    conn = pymysql.connect(**DB)
    cursor = conn.cursor()
    inserted = updated = 0

    for rec in registros:
        campos = [k for k, v in rec.items() if v is not None]
        valores = [rec[c] for c in campos]
        ph = ', '.join(['%s'] * len(campos))
        cols = ', '.join(campos)
        upd = ', '.join(
            f"{c} = VALUES({c})" for c in campos
            if c not in ('fecha', 'turno')
        )
        sql = f"""
            INSERT INTO condiciones_ro_turno ({cols})
            VALUES ({ph})
            ON DUPLICATE KEY UPDATE {upd}
        """
        cursor.execute(sql, valores)
        if cursor.rowcount == 1:
            inserted += 1
        elif cursor.rowcount == 2:
            updated += 1

    conn.commit()
    cursor.close()
    conn.close()

    print("=" * 60)
    print(f"  Insertados:   {inserted}")
    print(f"  Actualizados: {updated}")
    print(f"  Total:        {inserted + updated}")
    print("=" * 60)


if __name__ == '__main__':
    main()
