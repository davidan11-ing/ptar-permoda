"""Loader Proyección Química: lee la hoja PROYECCIÓN INSUMOS QUÍMICOS
del archivo PROYECCIÓN COSTOS 2026 ... y carga:

  - proyeccion_caudal_mensual    (caudal Plan Maestro por mes y subsistema)
  - proyeccion_quimica_mensual   (12 meses × kg/$ proyectado por químico)

Estructura del Excel:
  Bloque GEM+RO (filas 5-23, sistema='GEM_RO' caudal):
    Fila 5     : caudal Plan Maestro por mes (cols 7..18 = ene..dic)
    Fila 6-20  : un químico por fila
                  col 3 = nombre
                  col 5 = dosificación kg/m³
                  cols 7..18 = kg proyectado por mes
                  col 20 = $/Kg
                  col 21 = $ total año

  Bloque PTAP (filas 28-35, sistema='PTAP'):
    Fila 28    : caudal Plan Maestro PTAP por mes
    Fila 29-34 : un químico por fila (mismo formato)

Uso:
  python scripts/loaders/load_proyeccion_quimica.py "PROYECCIÓN COSTOS 2026 ...xlsx"
"""
from __future__ import annotations
import sys, argparse, shutil, tempfile
from datetime import date
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from db import get_engine  # noqa: E402

HOJA = "PROYECCIÓN INSUMOS QUÍMICOS"

# Filas con la fila CAUDAL PLAN MAESTRO de cada bloque
FILA_CAUDAL_GEM_RO = 5
FILA_CAUDAL_PTAP   = 28

# Rangos de filas de químicos
QUIMICOS_GEM_RO = (6, 21)   # incluye 6, excluye 21 (filas 6..20)
QUIMICOS_PTAP   = (29, 35)  # filas 29..34

# Columnas
COL_NOMBRE      = 3
COL_DOSIFIC     = 5    # kg/m3
COL_PRECIO_KG   = 20
COL_MES_INICIO  = 7    # enero
COL_MES_FIN     = 18   # diciembre

# Mapeo nombre Excel → (catálogo nombre, sistema-en-cada-bloque)
# El sistema depende del BLOQUE en que aparezca el químico:
#   bloque GEM+RO: usar la columna "sistema_gem_ro" de este mapeo
#   bloque PTAP:   siempre sistema='PTAP'
PROY_TO_PRODUCTO = {
    # GEM+RO chemicals - mapean a productos del catálogo y sistema GEM o RO
    "ACIDO CLORHIDRICO 31,5%":                                     ("HCL 10%",            "RO"),
    "SODIO BISULFITO":                                             ("BISULFITO DE SODIO", "RO"),
    "NITRATO DE PLATA":                                            ("NITRATO DE PLATA",   "OTRO"),
    "ACIDIFICANTE":                                                ("ACIDO",              "GEM"),
    "QUIMICO FLOCULANTE PARA TRATAMIENTO DE AGUA (Catiónico)":     ("POLIMERO CATIONICO", "GEM"),
    "QUIMICO FLOCULANTE PARA TRATAMIENTO DE AGUA (Aniónico)":      ("POLIMERO ANIONICO",  "GEM"),
    "QUIMICO COAGULANTE PARA TRATAMIENTO DE AGUA (Coagulante)":    ("COAGULANTE",         "GEM"),
    "QUIMICO COAGULANTE PARA TRATAMIENTO DE AGUA (Decolorante)":   ("DECOLORANTE",        "GEM"),
    "HIPOCLORITO DE SODIO":                                        ("HIPOCLORITO SODIO",  "OTRO"),
    "SODA CAUSTICA LIQUI48%  90 KL":                               ("SODA CAUSTICA",      "OTRO"),
    "QUIMICO ANTINCRUSTANTE PARA ÓSMOSIS INVERSA":                 ("VITEC 7000",         "RO"),
    "QUIMICO PARA TRATAMIENTO CIP ALCALINO ÓSMOSIS INVERSA":       ("CIP ALCALINO RO",    "RO"),
    "QUIMICO PARA TRATAMIENTO BIODISPERSANTE ÓSMOSIS INVERSA":     ("KURIVERTER IK-220",  "RO"),
    "QUIMICO PARA TRATAMIENTO CIP ACIDO ÓSMOSIS INVERSA":          ("CIP ACIDO RO",       "RO"),
    "ACIDO CÍTRICO":                                               ("ACIDO CITRICO",      "RO"),
}


def num(v) -> float | None:
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(',', '.')
        if not s or s.startswith('#'): return None
        try: return float(s)
        except ValueError: return None
    return None


def cargar_catalogo_productos(engine) -> dict[str, int]:
    """{nombre -> id} de producto_quimico."""
    with engine.connect() as conn:
        return {r[0]: r[1] for r in conn.execute(text(
            "SELECT nombre, id FROM producto_quimico"
        )).all()}


def procesar_caudal(ws, fila: int, sistema: str, anio: int) -> list[dict]:
    """Lee una fila de caudal y devuelve [{anio,mes,sistema,caudal_m3}, ...]"""
    out = []
    for mes_idx, col in enumerate(range(COL_MES_INICIO, COL_MES_FIN + 1), start=1):
        v = num(ws.cell(fila, col).value)
        if v is None or v == 0:
            continue
        out.append({
            "anio": anio, "mes": mes_idx, "sistema": sistema,
            "caudal_m3": round(v, 2),
        })
    return out


def procesar_quimicos(ws, fila_ini: int, fila_fin: int, bloque_es_ptap: bool,
                       productos: dict[str, int], anio: int) -> tuple[list[dict], list[str]]:
    """Lee filas de químicos y devuelve (registros, no_mapeados)."""
    registros = []
    no_mapeados = []

    for r in range(fila_ini, fila_fin):
        nombre_excel = ws.cell(r, COL_NOMBRE).value
        if not nombre_excel:
            continue
        nombre_excel = str(nombre_excel).strip()
        if nombre_excel.upper().startswith("SUBTOTAL"):
            continue

        if nombre_excel not in PROY_TO_PRODUCTO:
            no_mapeados.append(nombre_excel)
            continue

        producto_nombre, sistema_gem_ro = PROY_TO_PRODUCTO[nombre_excel]
        sistema = "PTAP" if bloque_es_ptap else sistema_gem_ro

        producto_id = productos.get(producto_nombre)
        if producto_id is None:
            no_mapeados.append(f"{nombre_excel} (catalogo no tiene '{producto_nombre}')")
            continue

        dosifi   = num(ws.cell(r, COL_DOSIFIC).value)
        precio   = num(ws.cell(r, COL_PRECIO_KG).value)

        for mes_idx, col in enumerate(range(COL_MES_INICIO, COL_MES_FIN + 1), start=1):
            kg = num(ws.cell(r, col).value)
            if kg is None or kg == 0:
                continue
            costo = round(kg * precio, 2) if (precio is not None) else None
            registros.append({
                "anio": anio, "mes": mes_idx,
                "producto_id": producto_id, "sistema": sistema,
                "dosificacion_kg_m3": (round(dosifi, 8) if dosifi is not None else None),
                "kg_proyectado": round(kg, 4),
                "costo_unitario_kg": (round(precio, 2) if precio is not None else None),
                "costo_proyectado": costo,
            })

    return registros, no_mapeados


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Ruta al archivo PROYECCION COSTOS 2026 ...xlsx")
    ap.add_argument("--anio", type=int, default=2026, help="Año de la proyección (default 2026)")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"[ERROR] No existe: {path}"); sys.exit(1)

    print(f"\n[PROYECCIÓN QUÍMICA] {path.name}  (año {args.anio})")

    tmp = Path(tempfile.mktemp(suffix=".xlsx")); shutil.copy2(path, tmp)
    try:
        wb = openpyxl.load_workbook(str(tmp), read_only=True, data_only=True)
        if HOJA not in wb.sheetnames:
            print(f"[ERROR] Hoja '{HOJA}' no encontrada en {wb.sheetnames}")
            sys.exit(1)
        ws = wb[HOJA]

        engine = get_engine()
        productos = cargar_catalogo_productos(engine)
        print(f"  Catálogo cargado: {len(productos)} productos")

        # 1) CAUDAL
        caudales = (
            procesar_caudal(ws, FILA_CAUDAL_GEM_RO, "GEM_RO", args.anio)
            + procesar_caudal(ws, FILA_CAUDAL_PTAP, "PTAP",   args.anio)
        )

        # 2) QUIMICOS — bloque GEM+RO
        regs_gemro, no_map_a = procesar_quimicos(
            ws, *QUIMICOS_GEM_RO, bloque_es_ptap=False,
            productos=productos, anio=args.anio,
        )
        # 3) QUIMICOS — bloque PTAP
        regs_ptap, no_map_b = procesar_quimicos(
            ws, *QUIMICOS_PTAP, bloque_es_ptap=True,
            productos=productos, anio=args.anio,
        )
        wb.close()

        registros = regs_gemro + regs_ptap
        no_mapeados = list(set(no_map_a + no_map_b))
        if no_mapeados:
            print(f"  [WARN] Sin mapear ({len(no_mapeados)}):")
            for n in no_mapeados: print(f"     - {n}")

        # 4) UPSERT en BD
        with engine.begin() as conn:
            # limpia el año en cuestión antes de cargar
            conn.execute(text("DELETE FROM proyeccion_caudal_mensual  WHERE anio = :a"),
                         {"a": args.anio})
            conn.execute(text("DELETE FROM proyeccion_quimica_mensual WHERE anio = :a"),
                         {"a": args.anio})

            if caudales:
                conn.execute(text("""
                    INSERT INTO proyeccion_caudal_mensual (anio, mes, sistema, caudal_m3)
                    VALUES (:anio, :mes, :sistema, :caudal_m3)
                """), caudales)

            if registros:
                conn.execute(text("""
                    INSERT INTO proyeccion_quimica_mensual
                      (anio, mes, producto_id, sistema, dosificacion_kg_m3,
                       kg_proyectado, costo_unitario_kg, costo_proyectado)
                    VALUES
                      (:anio, :mes, :producto_id, :sistema, :dosificacion_kg_m3,
                       :kg_proyectado, :costo_unitario_kg, :costo_proyectado)
                """), registros)

        print(f"\n  Caudales cargados:    {len(caudales)} filas")
        print(f"  Químicos cargados:    {len(registros)} filas")

        # Resumen
        with engine.connect() as conn:
            st = conn.execute(text("""
                SELECT
                    (SELECT COUNT(*) FROM proyeccion_caudal_mensual)  AS caud,
                    (SELECT COUNT(*) FROM proyeccion_quimica_mensual) AS quim,
                    (SELECT COUNT(DISTINCT producto_id) FROM proyeccion_quimica_mensual) AS prod,
                    (SELECT ROUND(SUM(kg_proyectado),0) FROM proyeccion_quimica_mensual) AS kg_total,
                    (SELECT ROUND(SUM(costo_proyectado),0) FROM proyeccion_quimica_mensual) AS costo_total
            """)).one()
        print(f"\n  En BD:")
        print(f"    Filas caudal:          {st[0]}")
        print(f"    Filas químicos:        {st[1]}")
        print(f"    Productos distintos:   {st[2]}")
        print(f"    Total kg proyectados:  {st[3]:,.0f}")
        print(f"    Total $ proyectados:   ${st[4]:,.0f}")
        print("DONE.")

    finally:
        tmp.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
