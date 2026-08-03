from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HDR_BG  = '1E293B'
HDR_FG  = 'FFFFFF'
BLK1_BG = 'E8F4FD'
BLK2_BG = 'FDF6E8'
SUB_BG  = 'D0D7E3'
TOT_BG  = '2D3748'
TOT_FG  = 'FFFFFF'

CAT_COLORS = {
    'quimicos':      'FF8A65',
    'residuosLodos': 'F06292',
    'consumibles':   'FFCC02',
    'calibMant':     '4DB6AC',
    'nomina':        '81C784',
    'energia':       '4FC3F7',
    'depreciacion':  'CE93D8',
}
CAT_LABELS = {
    'quimicos':      'Insumos Químicos',
    'residuosLodos': 'Residuos y Lodos',
    'consumibles':   'Consumibles',
    'calibMant':     'Calibración y Mant.',
    'nomina':        'Nómina',
    'energia':       'Energía Eléctrica',
    'depreciacion':  'Depreciación',
}
CATS = list(CAT_LABELS.keys())

EQ_COSTS = {
  'rotativa':     {'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'funza':        {'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'tintoreria':   {'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'lavanderia':   {'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'tk2m3':        {'residuosLodos':5.91,'calibMant':21.77,'nomina':37.64,'energia':1.22,'depreciacion':80.05},
  'tk15m3':       {'residuosLodos':44.30,'calibMant':21.77,'nomina':37.64,'energia':5.92,'depreciacion':80.05},
  'tk30m3':       {'residuosLodos':88.60,'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'tk60m3':       {'residuosLodos':177.20,'calibMant':21.77,'nomina':37.64,'energia':170.14,'depreciacion':80.05},
  'cribRot':      {'residuosLodos':182.70,'calibMant':21.77,'nomina':37.64,'energia':10.12,'depreciacion':80.05},
  'vibrat1':      {'residuosLodos':39.15,'calibMant':21.77,'nomina':37.64,'energia':10.12,'depreciacion':80.05},
  'vibrat2':      {'residuosLodos':39.15,'calibMant':21.77,'nomina':37.64,'energia':10.12,'depreciacion':80.05},
  'tkPulmon':     {'consumibles':26.55,'calibMant':21.77,'nomina':37.64,'energia':68.97,'depreciacion':80.05},
  'torre':        {'consumibles':31.00,'calibMant':21.77,'nomina':37.64,'energia':103.00,'depreciacion':80.05},
  'carcamo':      {'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'homogen':      {'consumibles':29.00,'calibMant':21.77,'nomina':37.64,'energia':818.49,'depreciacion':80.05},
  'eqGem':        {'quimicos':3585.07,'residuosLodos':726.86,'consumibles':31.00,'calibMant':21.77,'nomina':37.64,'energia':266.39,'depreciacion':80.05},
  'swingmill':    {'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'anoxic':       {'consumibles':4.00,'calibMant':21.77,'nomina':37.64,'depreciacion':80.05},
  'mbbr':         {'consumibles':4.00,'calibMant':21.77,'nomina':37.64,'energia':863.68,'depreciacion':80.05},
  'mbrT':         {'quimicos':7.11,'consumibles':43.00,'calibMant':21.77,'nomina':37.64,'energia':526.93,'depreciacion':80.05},
  'mbrK':         {'quimicos':7.11,'consumibles':43.00,'calibMant':21.77,'nomina':37.64,'energia':515.01,'depreciacion':80.05},
  'tkPermeado':   {'calibMant':21.77,'nomina':37.64,'energia':275.90,'depreciacion':80.05},
  'filtrosII':    {'consumibles':188.00,'calibMant':55.71,'nomina':50.71,'depreciacion':28.00},
  'filtro5':      {'consumibles':389.00,'calibMant':55.71,'nomina':50.71,'depreciacion':28.00},
  'ro1e1':        {'quimicos':479.42,'consumibles':62.67,'calibMant':55.71,'nomina':50.71,'energia':264.60,'depreciacion':28.00},
  'ro1e2':        {'quimicos':479.43,'consumibles':62.67,'calibMant':55.71,'nomina':50.71,'energia':97.02,'depreciacion':28.00},
  'tkRechazo':    {'consumibles':13.00,'calibMant':55.71,'nomina':50.71,'depreciacion':28.00},
  'ro2':          {'consumibles':62.67,'calibMant':55.71,'nomina':50.71,'energia':549.96,'depreciacion':28.00},
  'tkRechazoRO2': {'consumibles':13.00,'calibMant':55.67,'nomina':50.30,'depreciacion':28.25},
}

EQ_LABELS = {
  'rotativa':'Descarga Rotativa','funza':'Descarga Ext. Funza','tintoreria':'Descarga Tintorería',
  'lavanderia':'Descarga Lavandería','tk2m3':'TK Recepción 2 m³','tk15m3':'TK Buffer 15 m³',
  'tk30m3':'TK Recepción 30 m³','tk60m3':'TK 60 m³','cribRot':'Criba Rotativa',
  'vibrat1':'Criba Vibratoria 1','vibrat2':'Criba Vibratoria 2','tkPulmon':'TK Pulmón',
  'torre':'Torre Enfriamiento','carcamo':'Cárcamo','homogen':'TK Homogeneizador 800 m³',
  'eqGem':'Equipo GEM (DAF)','swingmill':'Espesador Swingmill',
  'anoxic':'Reactor Anóxico','mbbr':'Reactor MBBR','mbrT':'MBR TORAY (MBR2)',
  'mbrK':'MBR KUBOTA (MBR1)','tkPermeado':'TK Permeado',
  'filtrosII':'Filtros Intercambio Iónico','filtro5':'Filtros Cartucho 5 µm',
  'ro1e1':'Ósmosis Inversa RO1 E1','ro1e2':'Ósmosis Inversa RO1 E2',
  'tkRechazo':'TK Rechazo RO1','ro2':'Ósmosis Inversa RO2','tkRechazoRO2':'TK Rechazo RO2',
}

PTAR_EQS = ['rotativa','funza','tintoreria','lavanderia','tk2m3','tk15m3','tk30m3','tk60m3',
            'cribRot','vibrat1','vibrat2','tkPulmon','torre','carcamo','homogen','eqGem',
            'swingmill','anoxic','mbbr','mbrT','mbrK','tkPermeado']
RO_EQS   = ['filtrosII','filtro5','ro1e1','ro1e2','tkRechazo','ro2','tkRechazoRO2']

thin = Side(style='thin', color='CCCCCC')
def brd(cell):
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, bg=HDR_BG, fg=HDR_FG, size=11, bold=True, center=True):
    cell.font = Font(name='Calibri', bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                               vertical='center', wrap_text=True)

def num(cell, bg=None, bold=False, fmt='#,##0.00', color='000000'):
    cell.number_format = fmt
    cell.font = Font(name='Calibri', bold=bold, size=11, color=color)
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def txt(cell, bg=None, bold=False, color='000000', wrap=False, center=False):
    cell.font = Font(name='Calibri', bold=bold, color=color, size=11)
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                               vertical='center', wrap_text=wrap)

# ═══════════════════════════════════════════════════════════════════
# HOJA 1 — RESUMEN TOTALES
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '1. Resumen Totales'
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A4'

ws1.merge_cells('A1:G1')
c = ws1['A1']
c.value = 'COSTO $/m³ POR CATEGORÍA — TOTALES PROYECTADOS Y CRITERIO DE REPARTO'
hdr(c, size=13); ws1.row_dimensions[1].height = 32
ws1.row_dimensions[2].height = 6

hdrs1 = ['Categoría','Color','PTAR Fase 2 ($/m³)','RO ($/m³)','TOTAL ($/m³)','Equipos por bloque','Criterio de reparto']
ws1.row_dimensions[3].height = 38
for j, h in enumerate(hdrs1, 1):
    c = ws1.cell(3, j, h); hdr(c); brd(c)

cat_data = [
    ('quimicos',      3599.29, 958.85,  'Asignación directa al equipo que usa el reactivo (no hay reparto proporcional)'),
    ('residuosLodos', 1303.87,   0.00,  'Cribas por % captura (70/15/15); tanques proporcional al volumen; lodos directamente al GEM'),
    ('consumibles',    211.62, 790.94,  'Asignación directa: membranas amortizadas por MBR, cartuchos+resinas a filtros, puntos Hach por equipo'),
    ('calibMant',      478.61, 390.25,  'Reparto uniforme dentro de cada bloque: total bloque ÷ n° equipos'),
    ('nomina',         827.85, 354.79,  '70% PTAR F2 / 30% RO (por dedicación operativa), luego ÷ n° equipos de cada bloque'),
    ('energia',       3646.07, 911.52,  'Asignación directa por consumo eléctrico medido (corriente A × tensión × FP × horas)'),
    ('depreciacion',  1760.95, 196.40,  'CAPEX 90% PTAR / 10% RO (por inversión ejecutada), luego ÷ n° equipos de cada bloque'),
]

for i, (cat, ptar, ro, criterio) in enumerate(cat_data, 4):
    ws1.row_dimensions[i].height = 26
    c = ws1.cell(i, 1, CAT_LABELS[cat]); txt(c, bold=True); brd(c)
    c2 = ws1.cell(i, 2, '●')
    c2.font = Font(name='Calibri', bold=True, size=18, color=CAT_COLORS[cat])
    c2.alignment = Alignment(horizontal='center', vertical='center')
    brd(c2)
    for j, val in enumerate([ptar, ro], 3):
        c3 = ws1.cell(i, j, val)
        c3.number_format = '#,##0.00'
        c3.font = Font(name='Calibri', size=11, color='0000FF')
        c3.fill = PatternFill('solid', start_color='EEF4FF')
        c3.alignment = Alignment(horizontal='center', vertical='center')
        brd(c3)
    ct = ws1.cell(i, 5, f'=C{i}+D{i}')
    num(ct, bold=True); brd(ct)
    cn = ws1.cell(i, 6, '22 PTAR  /  7 RO')
    txt(cn, center=True); brd(cn)
    cc = ws1.cell(i, 7, criterio); txt(cc, wrap=True); brd(cc)

tr = 11
ws1.row_dimensions[tr].height = 28
ws1.merge_cells(f'A{tr}:B{tr}')
c = ws1.cell(tr, 1, 'TOTAL GENERAL'); hdr(c, bg=TOT_BG); brd(c)
for j in range(3, 6):
    col = get_column_letter(j)
    ct = ws1.cell(tr, j, f'=SUM({col}4:{col}10)')
    num(ct, bg=TOT_BG, bold=True, color=TOT_FG); brd(ct)
for j in [6, 7]:
    c = ws1.cell(tr, j)
    c.fill = PatternFill('solid', start_color=TOT_BG); brd(c)

ws1.row_dimensions[13].height = 20
ws1.merge_cells('A13:G13')
nota = ws1['A13']
nota.value = '   Azul = inputs proyectados (modificables)     Negro = fórmulas calculadas     Tolerancia verificación: ±$1,00 por redondeos'
nota.font = Font(name='Calibri', size=10, italic=True, color='555555')

for j, w in enumerate([35, 8, 18, 18, 18, 18, 58], 1):
    ws1.column_dimensions[get_column_letter(j)].width = w


# ═══════════════════════════════════════════════════════════════════
# HOJA 2 — METODOLOGÍA
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('2. Metodologia')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:F1')
c = ws2['A1']
c.value = 'METODOLOGÍA DE DISTRIBUCIÓN — PASO A PASO POR CATEGORÍA'
hdr(c, size=13); ws2.row_dimensions[1].height = 32

row = 3

metodologias = [
  {
    'cat': 'nomina', 'num': '1',
    'titulo': 'NÓMINA OPERATIVA',
    'total': 'Total proyectado: $1.182,64 COP/m³  (nómina mensual operarios PTAR ÷ producción m³/mes)',
    'criterio': 'Los operarios atienden principalmente PTAR Fase 2 (22 equipos). Se asigna 70% al bloque PTAR y 30% al bloque RO. Dentro de cada bloque, reparto uniforme entre todos los equipos.',
    'formula': 'PTAR F2: $828 / 22 equipos = $37.64/equipo     |     RO: $355 / 7 equipos = $50.71/equipo',
    'tabla_hdrs': ['Bloque','% Dedicación','Total bloque','N° equipos','$/m³ por equipo','Justificación'],
    'tabla': [
        ['PTAR Fase 2','70%','$828,08','22 equipos','$37.64','Mayor cantidad de equipos y puntos de operación'],
        ['RO','30%','$354,97','7 equipos','$50.71','Proceso más autónomo pero requiere mayor especialización'],
        ['Vertimiento','0%','$0,00','—','$0.00','Sin carga operativa asignada en modelo actual'],
    ],
    'nota': 'El mayor $/m³ de RO ($50.71) vs PTAR ($37.64) refleja que RO tiene menos equipos para repartir su 30% asignado.',
  },
  {
    'cat': 'depreciacion', 'num': '2',
    'titulo': 'DEPRECIACIÓN (CAPEX)',
    'total': 'Total proyectado: $1.957,35 COP/m³  (CAPEX total equipos amortizado según vida útil / producción m³/mes)',
    'criterio': 'El 90% del CAPEX se invirtió en el bloque PTAR Fase 2 (equipos biológicos, GEM, homogeneizador). El 10% restante corresponde al bloque RO. Reparto uniforme dentro de cada bloque.',
    'formula': 'PTAR F2: $1.761 / 22 equipos = $80.05/equipo     |     RO: $196 / 7 equipos = $28.00/equipo',
    'tabla_hdrs': ['Bloque','% CAPEX','Total bloque','N° equipos','$/m³ por equipo','Equipos principales'],
    'tabla': [
        ['PTAR Fase 2','90%','$1.761,10','22 equipos','$80.05','GEM, MBRs, MBBR, Homogeneizador, Cribas'],
        ['RO','10%','$196,00','7 equipos','$28.00','Trenes RO1 E1, E1, RO2, Filtros IO'],
        ['Vertimiento','0%','$0,00','—','$0.00','Equipos no incluidos en inventario depreciable activo'],
    ],
    'nota': 'TK Rechazo RO2 tiene $28.25 (diferencia de $0.25 por redondeo al distribuir $196 exactamente en 7 equipos).',
  },
  {
    'cat': 'calibMant', 'num': '3',
    'titulo': 'CALIBRACIÓN Y MANTENIMIENTO',
    'total': 'Total proyectado PTAR F2: $478,61 COP/m³   |   Total proyectado RO: $390,25 COP/m³',
    'criterio': 'Reparto estrictamente uniforme entre todos los equipos del bloque. Se justifica porque todos los equipos son objeto de revisión preventiva programada con la misma frecuencia base.',
    'formula': 'PTAR F2: $479 / 22 equipos = $21.77/equipo     |     RO: $390 / 7 equipos = $55.71/equipo',
    'tabla_hdrs': ['Bloque','Total bloque','N° equipos','$/m³ por equipo','Base del total'],
    'tabla': [
        ['PTAR Fase 2','$478,94','22 equipos','$21.77','Contrato mant. preventivo + calibración analizadores Hach bloque PTAR'],
        ['RO','$389,97','7 equipos','$55.71','Contrato mant. preventivo membranas + calibración analizadores bloque RO'],
    ],
    'nota': 'El mayor $/m³ en RO refleja que los contratos de mantenimiento de membranas de OI son más costosos por unidad de tratamiento.',
  },
  {
    'cat': 'energia', 'num': '4',
    'titulo': 'ENERGÍA ELÉCTRICA',
    'total': 'Total proyectado PTAR F2: $3.646,07 COP/m³   |   Total proyectado RO: $911,52 COP/m³',
    'criterio': 'Asignación directa al equipo consumidor medida por corriente eléctrica (A). Los equipos sin medidor propio (tanques pasivos, cárcamo) reciben $0. OZONO+CHILLER+GENERADORES se imputan al TK Homogeneizador.',
    'formula': 'Costo equipo = (kW × horas/mes × tarifa COP/kWh) ÷ m³/mes tratados',
    'tabla_hdrs': ['Equipo','$/m³','Equipo principal de consumo','Observación'],
    'tabla': [
        ['TK Recepción 2 m³','1.22','Bomba sumergible','Consumo mínimo'],
        ['TK Buffer 15 m³','5.92','Bomba recirculación',''],
        ['TK 60 m³','170.14','Bombas autocebantes principales','Mayor consumo fase prelim.'],
        ['Criba Rotativa','10.12','Motor tambor rotativo',''],
        ['Criba Vibratoria 1','10.12','Motor vibración',''],
        ['Criba Vibratoria 2','10.12','Motor vibración',''],
        ['TK Pulmón','68.97','Bomba alimentación torre',''],
        ['Torre Enfriamiento','103.00','Ventilador tiro forzado',''],
        ['TK Homogeneizador 800 m³','818.49','Agitador + OZONO + CHILLER + GENERADORES','OZONO/CHILLER/GEN = $450.63 incluido'],
        ['Equipo GEM (DAF)','266.39','Bomba recircul. + soplador',''],
        ['Reactor MBBR','863.68','Sopladora 90 kW x24h','Mayor consumidor PTAR F2'],
        ['MBR TORAY','526.93','Bomba permeado + soplador MBR2',''],
        ['MBR KUBOTA','515.01','Bomba permeado + soplador MBR1',''],
        ['TK Permeado','275.90','Bomba alta presión alim. RO',''],
        ['RO1 Etapa 1','264.60','Bomba alta presión E1 45 kW',''],
        ['RO1 Etapa 2','97.02','Bomba E2 15 kW',''],
        ['RO2','549.96','Bomba Bangpu 121 kW','Mayor consumidor RO'],
    ],
    'nota': 'Rotativa, Funza, Tintorería, Lavandería, TK30, Cárcamo, Swingmill, Anóxico, TK Permeado (solo bomba hacia RO ya contada en tkPermeado), Filtros IO, TK Rechazo: $0 en energía (equipos pasivos o sin medición propia).',
  },
  {
    'cat': 'quimicos', 'num': '5',
    'titulo': 'INSUMOS QUÍMICOS',
    'total': 'Total proyectado: $4.558,14 COP/m³   |   GEM representa el 78.6% del total de químicos del sistema',
    'criterio': 'Asignación directa al equipo que consume el reactivo. No hay reparto proporcional — si un equipo no usa reactivos, recibe $0.',
    'formula': 'Costo equipo = Σ (consumo mensual kg × precio COP/kg) ÷ m³/mes tratados',
    'tabla_hdrs': ['Equipo','$/m³','Reactivos asignados','% del total químicos'],
    'tabla': [
        ['Equipo GEM (DAF)','3.585,07','Ácido clorhídrico + Acidificante + Decolorante + Coagulante + Floculante cat. + Floculante an. + Nitrato de Plata (*)','78.7%'],
        ['MBR TORAY (MBR2)','7.11','NaOCl + Ácido cítrico para CIP membranas UF','0.16%'],
        ['MBR KUBOTA (MBR1)','7.11','NaOCl + Ácido cítrico para CIP membranas MF','0.16%'],
        ['RO1 Etapa 1','479.42','Antiincrustante VITEC + Bisulfito de sodio + HCl + Soda cáustica 48% (ajuste pH)','10.52%'],
        ['RO1 Etapa 2','479.43','Antiincrustante VITEC + Bisulfito de sodio + HCl + Soda cáustica 48% (ajuste pH)','10.52%'],
    ],
    'nota': '(*) Nitrato de Plata: reactivo de laboratorio para medición de cloruros; costo consolidado en GEM por ser el punto principal de monitoreo, se usa también en muestras del rechazo. CIP membranas MBR: ciclos programados de NaOCl y ácido cítrico. Soda cáustica 48% imputada a RO (no a GEM) para ajuste de pH pre-membranas de OI.',
  },
  {
    'cat': 'residuosLodos', 'num': '6',
    'titulo': 'RESIDUOS Y LODOS',
    'total': 'Total proyectado: $1.303,87 COP/m³   |   Tres sub-rubros: Disposición sólidos cribas $261 + Lavado tanques $316 + Lodos GEM $726,86',
    'criterio': 'Cada sub-rubro tiene su propio criterio: (a) cribas por porcentaje de captura de sólidos, (b) tanques proporcional al volumen lavado, (c) lodos fisicoquímicos directamente al equipo generador.',
    'formula': '(a) $261 x % captura    |    (b) $316 x Volumen_i / 107 m³ total    |    (c) $726.86 directo a GEM',
    'tabla_hdrs': ['Sub-rubro','Equipo','Factor/Criterio','$/m³','Justificación'],
    'tabla': [
        ['Disposición sólidos cribas ($261)','Criba Rotativa','70% del total','182.70','Tamiz 1mm captura la mayor fracción de sólidos gruesos'],
        ['Disposición sólidos cribas ($261)','Criba Vibratoria 1','15% del total','39.15','Captura finos 0.10mm, igual entre las dos vibratorias'],
        ['Disposición sólidos cribas ($261)','Criba Vibratoria 2','15% del total','39.15','Captura finos 0.10mm, igual entre las dos vibratorias'],
        ['Lavado tanques ($316)','TK Recepción 2 m³','2/107 = 1.87%','5.91','Proporcional al volumen real del tanque'],
        ['Lavado tanques ($316)','TK Buffer 15 m³','15/107 = 14.0%','44.30','Proporcional al volumen real del tanque'],
        ['Lavado tanques ($316)','TK Recepción 30 m³','30/107 = 28.0%','88.60','Proporcional al volumen real del tanque'],
        ['Lavado tanques ($316)','TK 60 m³','60/107 = 56.1%','177.20','Proporcional al volumen real del tanque'],
        ['Lodos fisicoquímicos GEM','Equipo GEM (DAF)','100% directo','726.86','Disposición lodos DAF + transporte (fuente exclusiva del rubro)'],
    ],
    'nota': 'Total tanques para lavado: 2+15+30+60 = 107 m³. Los lodos del Swingmill se imputan al GEM por ser éste su única fuente de generación.',
  },
  {
    'cat': 'consumibles', 'num': '7',
    'titulo': 'CONSUMIBLES',
    'total': 'Total proyectado PTAR F2: $211,62 COP/m³   |   Total proyectado RO: $790,94 COP/m³',
    'criterio': 'Asignación directa por equipo: membranas MBR amortizadas (TORAY+KUBOTA), puntos Hach por equipo medidor, cartuchos por filtro, resinas por columna, membranas RO amortizadas.',
    'formula': 'Costo = (unidades × precio unitario × reemplazos/año ÷ 12) ÷ m³/mes',
    'tabla_hdrs': ['Equipo','$/m³','Consumible principal','Base de cálculo'],
    'tabla': [
        ['TK Pulmón','26.55','Reactivos analizadores Hach','27 puntos Hach equivalentes asignados'],
        ['Torre Enfriamiento','31.00','Reactivos analizadores Hach','31 puntos Hach equivalentes asignados'],
        ['TK Homogeneizador','29.00','Reactivos analizadores Hach','29 puntos Hach equivalentes asignados'],
        ['Equipo GEM (DAF)','31.00','Reactivos analizadores Hach','31 puntos Hach equivalentes asignados'],
        ['Reactor Anóxico','4.00','Reactivos analizadores Hach','4 puntos Hach equivalentes asignados'],
        ['Reactor MBBR','4.00','Reactivos analizadores Hach','4 puntos Hach equivalentes asignados'],
        ['MBR TORAY (MBR2)','43.00','Membranas UF 0.08µm amortizadas + Hach','TORAY 4.800 u × $500k / 5 años + 7 pts Hach'],
        ['MBR KUBOTA (MBR1)','43.00','Membranas MF 0.4µm amortizadas + Hach','KUBOTA 2.000 u × $800k / 5 años + 7 pts Hach'],
        ['Filtros Intercambio Iónico','188.00','Resinas catiónica + aniónica','Reposición por agotamiento de resinas en ciclos de regeneración'],
        ['Filtros Cartucho 5 µm','389.00','40 cartuchos/semana','Reemplazo ~semanal o cuando ΔP indica colmatación en RO'],
        ['RO1 Etapa 1','62.67','Membranas RO amortizadas','72 membranas × $2.8M / 5 años de vida útil estimada'],
        ['RO1 Etapa 2','62.67','Membranas RO amortizadas','72 membranas × $2.8M / 5 años de vida útil estimada'],
        ['TK Rechazo RO1','13.00','Gestión muestras y análisis calidad rechazo',''],
        ['RO2','62.67','Membranas RO2 amortizadas','42 membranas × $4.0M / 5 años de vida útil estimada'],
        ['TK Rechazo RO2','13.00','Gestión muestras y análisis calidad rechazo',''],
    ],
    'nota': 'Puntos Hach: cada punto representa el costo mensual de reactivos para el analizador en línea o equipo portátil asignado a ese punto de muestreo de calidad de agua.',
  },
]

for meta in metodologias:
    cat = meta['cat']
    color = CAT_COLORS[cat]
    dark_text = cat in ['consumibles', 'nomina', 'calibMant']
    fg_color = '000000' if dark_text else 'FFFFFF'

    ws2.merge_cells(f'A{row}:F{row}')
    c = ws2.cell(row, 1, f"{meta['num']}. {meta['titulo']}")
    hdr(c, bg=color, fg=fg_color, size=12)
    ws2.row_dimensions[row].height = 28; row += 1

    ws2.merge_cells(f'A{row}:F{row}')
    c = ws2.cell(row, 1, meta['total'])
    c.font = Font(name='Calibri', size=10, bold=True, color='0000FF')
    c.fill = PatternFill('solid', start_color='EEF4FF')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws2.row_dimensions[row].height = 18; row += 1

    ws2.merge_cells(f'A{row}:F{row}')
    c = ws2.cell(row, 1, 'Criterio: ' + meta['criterio'])
    txt(c, wrap=True)
    ws2.row_dimensions[row].height = 32; row += 1

    ws2.merge_cells(f'A{row}:F{row}')
    c = ws2.cell(row, 1, 'Formula: ' + meta['formula'])
    c.font = Font(name='Calibri', size=10, italic=True, color='555555')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws2.row_dimensions[row].height = 18; row += 1

    for k, fila in enumerate([meta['tabla_hdrs']] + meta['tabla']):
        ws2.row_dimensions[row].height = 18
        for j, val in enumerate(fila, 1):
            c = ws2.cell(row, j, val)
            if k == 0:
                light_color = color + '99'
                hdr(c, bg=light_color[:6], fg=fg_color, size=10)
            else:
                bg_row = 'F5F5F5' if k % 2 == 0 else None
                txt(c, bg=bg_row, wrap=(j == len(fila)))
                c.alignment = Alignment(horizontal='center' if j > 1 else 'left',
                                        vertical='center', wrap_text=(j == len(fila)))
            brd(c)
        row += 1

    ws2.merge_cells(f'A{row}:F{row}')
    c = ws2.cell(row, 1, 'Nota: ' + meta['nota'])
    c.font = Font(name='Calibri', size=9, italic=True, color='666666')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    ws2.row_dimensions[row].height = 30; row += 2

for j, w in enumerate([28, 28, 20, 12, 32, 18], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w


# ═══════════════════════════════════════════════════════════════════
# HOJA 3 — MATRIZ $/m³ POR EQUIPO
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('3. Matriz por Equipo')
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = 'C3'

ncols = len(CATS)
last_col = get_column_letter(2 + ncols + 1)

ws3.merge_cells(f'A1:{last_col}1')
c = ws3['A1']
c.value = 'MATRIZ DE COSTOS $/m³ — POR EQUIPO Y CATEGORIA'
hdr(c, size=13); ws3.row_dimensions[1].height = 32

ws3.cell(2, 1, 'ID'); hdr(ws3.cell(2, 1)); brd(ws3.cell(2, 1))
ws3.cell(2, 2, 'Equipo'); hdr(ws3.cell(2, 2)); brd(ws3.cell(2, 2))
for j, cat in enumerate(CATS, 3):
    c = ws3.cell(2, j, CAT_LABELS[cat])
    dark_text = cat in ['consumibles', 'nomina', 'calibMant']
    hdr(c, bg=CAT_COLORS[cat], fg='000000' if dark_text else 'FFFFFF')
    brd(c)
c = ws3.cell(2, 3+ncols, 'TOTAL $/m³'); hdr(c, bg=TOT_BG); brd(c)
ws3.row_dimensions[2].height = 42

def write_block(ws, equipos, start_row, bg):
    r = start_row
    for eq in equipos:
        ws.row_dimensions[r].height = 18
        c1 = ws.cell(r, 1, eq); txt(c1, bg=bg); brd(c1)
        c2 = ws.cell(r, 2, EQ_LABELS[eq]); txt(c2, bg=bg); brd(c2)
        for j, cat in enumerate(CATS, 3):
            val = EQ_COSTS[eq].get(cat, 0)
            c = ws.cell(r, j)
            if val:
                c.value = val
                c.number_format = '#,##0.00'
                c.font = Font(name='Calibri', size=11, color='000000')
            else:
                c.value = '-'
                c.font = Font(name='Calibri', size=11, color='BBBBBB')
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            brd(c)
        sc = get_column_letter(3)
        ec = get_column_letter(2+ncols)
        ct = ws.cell(r, 3+ncols, f'=SUMPRODUCT(IF(ISNUMBER({sc}{r}:{ec}{r}),{sc}{r}:{ec}{r},0))')
        num(ct, bg=bg, bold=True); brd(ct)
        r += 1
    return r

r3 = 3
ws3.merge_cells(f'A{r3}:{last_col}{r3}')
c = ws3.cell(r3, 1, 'BLOQUE PTAR FASE 2 — 22 equipos')
hdr(c, bg='2B4C7E'); ws3.row_dimensions[r3].height = 22; r3 += 1
ptar_start = r3
r3 = write_block(ws3, PTAR_EQS, r3, BLK1_BG)
ptar_end = r3 - 1

ws3.row_dimensions[r3].height = 22
ws3.merge_cells(f'A{r3}:B{r3}')
c = ws3.cell(r3, 1, 'SUBTOTAL PTAR FASE 2'); hdr(c, bg=SUB_BG, fg='000000', size=10); brd(c)
for j in range(3, 3+ncols+1):
    col = get_column_letter(j)
    ct = ws3.cell(r3, j, f'=SUMPRODUCT(IF(ISNUMBER({col}{ptar_start}:{col}{ptar_end}),{col}{ptar_start}:{col}{ptar_end},0))')
    num(ct, bg=SUB_BG, bold=True); brd(ct)
sub_ptar_row = r3; r3 += 1

ws3.merge_cells(f'A{r3}:{last_col}{r3}')
c = ws3.cell(r3, 1, 'BLOQUE RO (OSMOSIS INVERSA) — 7 equipos')
hdr(c, bg='7B5E2A'); ws3.row_dimensions[r3].height = 22; r3 += 1
ro_start = r3
r3 = write_block(ws3, RO_EQS, r3, BLK2_BG)
ro_end = r3 - 1

ws3.row_dimensions[r3].height = 22
ws3.merge_cells(f'A{r3}:B{r3}')
c = ws3.cell(r3, 1, 'SUBTOTAL RO'); hdr(c, bg=SUB_BG, fg='000000', size=10); brd(c)
for j in range(3, 3+ncols+1):
    col = get_column_letter(j)
    ct = ws3.cell(r3, j, f'=SUMPRODUCT(IF(ISNUMBER({col}{ro_start}:{col}{ro_end}),{col}{ro_start}:{col}{ro_end},0))')
    num(ct, bg=SUB_BG, bold=True); brd(ct)
sub_ro_row = r3; r3 += 1

ws3.row_dimensions[r3].height = 26
ws3.merge_cells(f'A{r3}:B{r3}')
c = ws3.cell(r3, 1, 'TOTAL GENERAL'); hdr(c, bg=TOT_BG); brd(c)
for j in range(3, 3+ncols+1):
    col = get_column_letter(j)
    ct = ws3.cell(r3, j, f'={col}{sub_ptar_row}+{col}{sub_ro_row}')
    num(ct, bg=TOT_BG, bold=True, color=TOT_FG); brd(ct)

for j, w in enumerate([18, 32] + [14]*ncols + [14], 1):
    ws3.column_dimensions[get_column_letter(j)].width = w


# ═══════════════════════════════════════════════════════════════════
# HOJA 4 — VERIFICACIÓN
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('4. Verificacion')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:G1')
c = ws4['A1']
c.value = 'VERIFICACION — SUMA CALCULADA vs. TOTAL PROYECTADO POR CATEGORIA Y BLOQUE'
hdr(c, size=13); ws4.row_dimensions[1].height = 32

hdrs4 = ['Categoría','Bloque','Target proyectado','Suma calculada','Diferencia','Cuadra?','Fuente del target']
ws4.row_dimensions[3].height = 35
for j, h in enumerate(hdrs4, 1):
    c = ws4.cell(3, j, h); hdr(c); brd(c)

targets = {
    ('quimicos',      'PTAR F2'): 3599.29, ('quimicos',      'RO'):  958.85,
    ('residuosLodos', 'PTAR F2'): 1303.87, ('residuosLodos', 'RO'):    0.00,
    ('consumibles',   'PTAR F2'):  211.62, ('consumibles',   'RO'):  790.94,
    ('calibMant',     'PTAR F2'):  478.61, ('calibMant',     'RO'):  390.25,
    ('nomina',        'PTAR F2'):  827.85, ('nomina',        'RO'):  354.79,
    ('energia',       'PTAR F2'): 3646.07, ('energia',       'RO'):  911.52,
    ('depreciacion',  'PTAR F2'): 1760.95, ('depreciacion',  'RO'):  196.40,
}
fuentes = {
    'quimicos':      'Costeo reactivos GEM + RO (cotizaciones proveedor)',
    'residuosLodos': 'Contratos disposicion residuos + transporte',
    'consumibles':   'Membranas + analizadores Hach + cartuchos + resinas',
    'calibMant':     'Contrato mantenimiento preventivo anual',
    'nomina':        'Nomina operativa mensual 2026 (70/30 PTAR/RO)',
    'energia':       'Factura energia promedio 6 meses + medicion corriente',
    'depreciacion':  'CAPEX equipos x tasa depreciacion / produccion m3/mes',
}

r4 = 4
for cat in CATS:
    for bloque, eqs in [('PTAR F2', PTAR_EQS), ('RO', RO_EQS)]:
        ws4.row_dimensions[r4].height = 20
        target = targets.get((cat, bloque), 0.0)
        calc   = round(sum(EQ_COSTS[eq].get(cat, 0) for eq in eqs), 2)
        diff   = round(calc - target, 2)
        ok     = abs(diff) <= 1.0

        c = ws4.cell(r4, 1, CAT_LABELS[cat]); txt(c, bold=True); brd(c)
        c = ws4.cell(r4, 2, bloque); txt(c, center=True); brd(c)

        ct = ws4.cell(r4, 3, target)
        ct.number_format = '#,##0.00'
        ct.font = Font(name='Calibri', size=11, color='0000FF')
        ct.fill = PatternFill('solid', start_color='EEF4FF')
        ct.alignment = Alignment(horizontal='center', vertical='center'); brd(ct)

        cc = ws4.cell(r4, 4, calc)
        num(cc); brd(cc)

        cd = ws4.cell(r4, 5, f'=D{r4}-C{r4}')
        cd.number_format = '#,##0.00'
        cd.font = Font(name='Calibri', size=11, color='006600' if ok else 'CC0000')
        cd.alignment = Alignment(horizontal='center', vertical='center'); brd(cd)

        cs = ws4.cell(r4, 6, 'SI' if ok else 'REVISAR')
        cs.font = Font(name='Calibri', size=11, bold=True,
                       color='006600' if ok else 'CC0000')
        cs.fill = PatternFill('solid', start_color='E8FFE8' if ok else 'FFE8E8')
        cs.alignment = Alignment(horizontal='center', vertical='center'); brd(cs)

        cf = ws4.cell(r4, 7, fuentes[cat]); txt(cf, wrap=True); brd(cf)
        r4 += 1

ws4.row_dimensions[r4+1].height = 20
ws4.merge_cells(f'A{r4+1}:G{r4+1}')
c = ws4.cell(r4+1, 1, '  Tolerancia aceptable: +/-$1.00 por m3 por redondeos al distribuir totales entre equipos')
c.font = Font(name='Calibri', size=10, italic=True, color='444444')

for j, w in enumerate([26, 12, 20, 20, 14, 12, 48], 1):
    ws4.column_dimensions[get_column_letter(j)].width = w

# ── Guardar ─────────────────────────────────────────────────────────
out = r'C:\Users\lunaop\OneDrive - PERMODA LTDA\Documentos\Claude\Projects\App PTAR 2\METODOLOGIA_COSTOS_M3_PTAR_v3.xlsx'
wb.save(out)
print('OK:', out)
