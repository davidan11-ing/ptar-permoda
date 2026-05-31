"""
load_calidad_excel.py
Lee la hoja "Tabla datos 1" del Excel DASHBOARD CALIDAD AGUA ABRIL 2026.xlsm
y carga los datos en medicion_calidad (sin alterar ninguna tabla).

Columnas de medicion_calidad: id, fecha, turno, parametro_id, unidad_id, valor, observacion
UNIQUE KEY uk_mc (fecha, turno, parametro_id, unidad_id)
"""

import sys
import os
import pymysql
from openpyxl import load_workbook
from datetime import datetime

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_PATH = os.getenv(
    "EXCEL_PATH",
    r"C:\Users\santi\OneDrive\Imágenes\Documentos\Claude\Ptar-Permoda\Bases de datos\DASHBOARD CALIDAD AGUA ABRIL 2026.xlsm"
)
SHEET_NAME = "Tabla datos 1"

MYSQL = dict(
    host=os.getenv("DB_HOST", "127.0.0.1"),
    port=int(os.getenv("DB_PORT", "3306")),
    user=os.getenv("DB_USER", "root"),
    password=os.getenv("DB_PASS", ""),   # set via env var or .env
    database=os.getenv("DB_NAME", "ptar_permoda"),
    charset="utf8mb4",
    autocommit=False,
)

# ── Mapeo Excel PARAMETRO (strip+upper) → parametro_id ───────────────────────
PARAM_MAP = {
    "TEMPERATURA(°C)":                                      1,
    "PH (UNIDADES DE PH)":                                  2,
    "DEMANDA QUÍMICA DE OXÍGENO (DQO)(MG/L)":               3,
    "SÓLIDOS SUSPENDIDOS TOTALES(MG/L)":                    4,
    "SÓLIDOS SEDIMENTABLES (MG/L)":                         5,
    "CLORUROS (MG/L)":                                      6,
    "CONDUCTIVIDAD(US/CM)":                                 7,
    "SOLIDOS DISUELTOS TOTALES (TDS)(MG/L)":                8,
    "COLOR (UPTCO)":                                        9,
    "SOLIDOS SUSPENDIDOS TOTALES GRAVIMETRICO(MG/L)":       10,
    "HIERRO(ML/L)":                                         11,
    "FOSFORO TOTAL(MG/L)":                                  12,
    "NITRÓGENO TOTAL(MG/L)":                                13,
    "SULFATOS (MG/L)":                                      14,
    "SILICE(MG/L)":                                         15,
    "ORP(-MV)":                                             16,
    "CLORO RES(MG/L)":                                      17,
    "TURBIDEZ(NTU)":                                        18,
    "ALCALINIDAD (MG CACO3/L)":                             19,
    "DUREZA TOTAL (MG CACO3/L)":                            20,
    "DUREZA CÁLCICA(MG CACO3/L)":                           21,
}

# ── Mapeo columna Excel (strip+upper) → unidad_id ────────────────────────────
# Las columnas de datos empiezan en la col índice 4 (0-based)
COL_UNIT_MAP = {
    "TANQUE PULMON":                        1,
    "TANQUE HOMOGENEIZADOR (ENTRADA GEM)":  2,
    "GEM (SALIDA)":                         3,
    "REACTOR ANÓXICO (INTERNO)":            4,
    "REACTOR MBBR (INTERNO)":               5,
    "REACTOR MBR 1 (INTERNO)":             6,
    "REACTOR MBR 2 (INTERNO)":             7,
    "REACTOR MBR 1 (PERMEADO)":            8,
    "REACTOR MBR 2 (PERMEADO)":            9,
    "VERTIMIENTO":                          10,
    "RO 1 (COMPUESTA - PERMEADO)":         11,
    "RO 1 (ETAPA 1)":                       12,
    "RO 1 (ETAPA 2)":                       13,
    "RO 2 (PERMEADO)":                      14,
    "RO (RECHAZO)":                         15,
}


def normalize(s):
    """Strip espacios, upper, normaliza tildes básicas para comparación."""
    if s is None:
        return ""
    return str(s).strip().upper()


def find_param_id(raw_param):
    key = normalize(raw_param)
    if key in PARAM_MAP:
        return PARAM_MAP[key]
    # Búsqueda parcial como fallback
    for k, v in PARAM_MAP.items():
        if key in k or k in key:
            return v
    return None


def main():
    print(f"Leyendo Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: hoja '{SHEET_NAME}' no encontrada. Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    print(f"Filas totales (con header): {len(rows)}")

    # Leer header (fila 0)
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    print(f"Columnas: {header[:6]}... (total {len(header)})")

    # Identificar índices de las columnas de datos (unidades)
    col_units = []  # lista de (col_index, unidad_id)
    for i, h in enumerate(header):
        uid = COL_UNIT_MAP.get(h.strip().upper())
        if uid is not None:
            col_units.append((i, uid))

    print(f"Columnas de unidad mapeadas: {len(col_units)}")
    if len(col_units) == 0:
        print("ERROR: ninguna columna mapeada. Revisa el header.")
        sys.exit(1)

    # Índices de columnas clave
    try:
        idx_fecha  = next(i for i, h in enumerate(header) if "FECHA" in h.upper())
        idx_param  = next(i for i, h in enumerate(header) if "PARAMETRO" in h.upper() or "PARÁMETRO" in h.upper())
        idx_turno  = next(i for i, h in enumerate(header) if h.strip().upper() == "TURNO")
    except StopIteration as e:
        print(f"ERROR buscando columnas clave: {e}")
        sys.exit(1)

    print(f"  idx_fecha={idx_fecha}, idx_param={idx_param}, idx_turno={idx_turno}")

    # ── Conectar MySQL ─────────────────────────────────────────────────────────
    conn = pymysql.connect(**MYSQL)
    cur  = conn.cursor()

    inserted = 0
    skipped  = 0
    no_param = 0
    no_value = 0

    INSERT_SQL = """
        INSERT IGNORE INTO medicion_calidad (fecha, turno, parametro_id, unidad_id, valor)
        VALUES (%s, %s, %s, %s, %s)
    """

    batch = []
    BATCH_SIZE = 500

    for row_num, row in enumerate(rows[1:], start=2):
        # Extraer campos clave
        raw_fecha  = row[idx_fecha]
        raw_param  = row[idx_param]
        raw_turno  = row[idx_turno]

        # Validar fecha
        if raw_fecha is None:
            continue
        if isinstance(raw_fecha, datetime):
            fecha = raw_fecha.date()
        else:
            try:
                fecha = datetime.strptime(str(raw_fecha)[:10], "%Y-%m-%d").date()
            except Exception:
                continue

        # Validar turno
        try:
            turno = int(raw_turno)
            if turno not in (1, 2, 3):
                continue
        except Exception:
            continue

        # Resolver parametro_id
        param_id = find_param_id(raw_param)
        if param_id is None:
            no_param += 1
            if no_param <= 5:
                print(f"  AVISO fila {row_num}: parámetro no mapeado → '{raw_param}'")
            continue

        # Insertar una fila por cada unidad con valor válido
        for col_idx, unit_id in col_units:
            val = row[col_idx] if col_idx < len(row) else None

            # Saltar nulos, vacíos y ceros exactos (sin dato real)
            if val is None or val == "" or val == 0:
                no_value += 1
                continue

            try:
                valor = float(val)
            except (TypeError, ValueError):
                no_value += 1
                continue

            batch.append((fecha, turno, param_id, unit_id, valor))

        # Flush batch
        if len(batch) >= BATCH_SIZE:
            cur.executemany(INSERT_SQL, batch)
            inserted += cur.rowcount
            conn.commit()
            batch.clear()
            print(f"  ... {inserted} filas insertadas hasta ahora ...", end="\r")

    # Último batch
    if batch:
        cur.executemany(INSERT_SQL, batch)
        inserted += cur.rowcount
        conn.commit()

    cur.close()
    conn.close()
    wb.close()

    print(f"\n{'='*50}")
    print(f"✅ Carga completada")
    print(f"   Filas insertadas  : {inserted}")
    print(f"   Sin parámetro     : {no_param}")
    print(f"   Valores nulos/cero: {no_value}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
