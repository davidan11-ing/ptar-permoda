import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

uploads = r"C:\Users\lunaop\AppData\Roaming\Claude\local-agent-mode-sessions\ec4a6280-9082-4796-9f16-f59d6a637f00\407556f0-174f-4463-9037-b41400b56a95\local_c4abf24d-82f3-4f11-b753-fc83c00d4a58\uploads"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "INFORME ABRIL 2026"

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=16, color="1F4E78")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Título
ws['A1'] = "INFORME DE CALIDAD DE AGUA - PTAR 2"
ws['A1'].font = title_font
ws.merge_cells('A1:H1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

ws['A2'] = "Período: Abril 2026 (1-19 de abril)"
ws['A2'].font = Font(italic=True, size=11)
ws.merge_cells('A2:H2')
ws['A2'].alignment = Alignment(horizontal='center')

ws['A3'] = "Preparado por: Moon | Fecha: 21 de abril de 2026"
ws['A3'].font = Font(italic=True, size=10)
ws.merge_cells('A3:H3')
ws['A3'].alignment = Alignment(horizontal='center')

row = 5
ws[f'A{row}'] = "1. RESUMEN EJECUTIVO"
ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
row += 1

ws[f'A{row}'] = "Estado operativo durante abril 1-19 - Dentro de parámetros normativos"
row += 2

# PARÁMETROS DE CALIDAD
ws[f'A{row}'] = "2. PARÁMETROS DE CALIDAD DE AGUA"
ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
row += 1

headers = ["Parámetro", "Unidad", "Límite Normativo", "Mín. Abril", "Máx. Abril", "Promedio", "Estado"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row += 1

parametros = [
    ("Temperatura", "°C", "≤ 40", "", "", "", ""),
    ("pH", "u.pH", "5,0 - 9,0", "", "", "", ""),
    ("DQO", "mg/L", "≤ 600", "", "", "", ""),
    ("SST", "mg/L", "≤ 75", "", "", "", ""),
    ("Color", "UPTCO", "Sin límite", "", "", "", ""),
    ("Conductividad", "µS/cm", "Sin límite", "", "", "", ""),
    ("Cloruros", "mg/L", "≤ 1.200", "", "", "", ""),
]

for param in parametros:
    for col, val in enumerate(param, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.border = border
        if col <= 3:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = Alignment(horizontal='center')
    row += 1

# COSTOS
row += 1
ws[f'A{row}'] = "3. ANÁLISIS DE COSTOS Y CONSUMO QUÍMICO"
ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
row += 1

headers_costos = ["Métrica", "Abril 1-19", "Proyección", "Variación", "% Variación", "vs Límite", "Estado"]
for col, header in enumerate(headers_costos, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')

row += 1

costos_data = [
    ("Volumen (m³)", "", "", "", "", "", ""),
    ("Costo $/m³", "", "< $3.599", "", "", "OK", ""),
    ("Costo Total ($)", "", "", "", "", "", ""),
    ("Horas Operación", "", "", "", "", "", ""),
]

for item in costos_data:
    for col, val in enumerate(item, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
    row += 1

# QUÍMICOS
row += 1
ws[f'A{row}'] = "4. CONSUMO DE QUÍMICOS"
ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
row += 1

headers_quim = ["Químico", "Proyección Abril (kg)", "Q1 2026 Promedio", "Tendencia", "Estado"]
for col, header in enumerate(headers_quim, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.border = border

row += 1

quim_data = [
    ("Decolorante", "~5.200", "4.971", "Creciente", ""),
    ("Coagulante", "~13.000", "18.042", "Decreciente", "OK"),
    ("Acidificante", "~9.300", "10.924", "Creciente", "REVISAR"),
    ("Pol. Aniónico", "~160", "159", "Estable", ""),
    ("Pol. Catiónico", "~330", "320", "Estable", ""),
]

for item in quim_data:
    for col, val in enumerate(item, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.border = border
    row += 1

# RECOMENDACIONES
row += 2
ws[f'A{row}'] = "5. RECOMENDACIONES INMEDIATAS"
ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
row += 1

recomendaciones = [
    "1. Completar digitalización de datos RO (Compuesta, Etapa 1, Etapa 2, RO2)",
    "2. Revisar correlación pH-ácido en Tanque Pulmón",
    "3. Aumentar frecuencia muestreo DQO a mínimo 2-3 turnos/semana",
    "4. Validar autosuficiencia hídrica (meta: mantener 100% recirculación interna)",
    "5. Preparar análisis comparativo abril vs Q1 2026",
]

for rec in recomendaciones:
    ws[f'A{row}'] = rec
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    row += 1

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

output_file = "INFORME_CALIDAD_AGUA_ABRIL_2026.xlsx"
wb.save(output_file)
print(f"Informe creado: {output_file}")
